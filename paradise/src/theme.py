# -*- coding: utf-8 -*-
"""Charte graphique du classeur PARADISE ALUMINIUM.

Les couleurs sont celles du logo et du dossier de presentation de la
societe : bleu marine, acier et or, sur des fonds tres clairs.
"""
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

FONT = 'Arial'

# Palette de la marque -------------------------------------------------
MARINE     = '021F47'   # bleu marine du logo   - bandeaux principaux
MARINE_2   = '0B2E5E'   # marine plus clair     - lignes de total
ACIER      = '3B5C8F'   # bleu acier            - sous-titres, en-tetes
OR         = 'C8A24B'   # or                    - filets et accents
ENCRE      = '1B2430'   # texte principal
GRIS       = '5B6675'   # texte secondaire
GRIS_CLAIR = '9AA3AE'   # texte tertiaire
BORDURE    = 'DDE2EA'   # filets discrets
FOND       = 'F3F5F9'   # fond clair des zones de lecture
FOND_2     = 'E8EDF5'   # fond clair alterne
BLANC      = 'FFFFFF'
VERT       = '1E6B3A'   # positif, encaissement
VERT_CLAIR = 'E4F1E8'
ROUGE      = 'A8231C'   # negatif, retard, depassement
ROUGE_CLAIR= 'FBE7E5'
AMBRE      = '8A5A00'   # alerte douce
AMBRE_CLAIR= 'FFF3D6'
BLEU_SAISIE= '1A56A0'   # texte des cellules a remplir
JAUNE      = 'FFF7D6'   # fond des cellules a remplir

# Formats de nombre ----------------------------------------------------
DH     = '#,##0 "DH";[Red]-#,##0 "DH";"–"'
DH2    = '#,##0.00 "DH";[Red]-#,##0.00 "DH";"–"'
DH_TOT = '#,##0 "DH";-#,##0 "DH";"–"'          # sur fond fonce : pas de rouge
NBR    = '#,##0;[Red]-#,##0;"–"'
ENT    = '0'                                    # année, numéro de mois
PCT    = '0.0%;[Red]-0.0%;"–"'
TEXTE  = '@'
DATE   = 'DD/MM/YYYY'
MOIS   = '[$-40C]MMMM YYYY'
JOUR_LONG = '[$-40C]dddd D MMMM YYYY'
JOUR_CRT  = '[$-40C]ddd DD/MM'


# Bordures -------------------------------------------------------------
def _s(color, style='thin'):
    return Side(style=style, color=color)


BORD_LEGER = Border(left=_s(BORDURE), right=_s(BORDURE),
                    top=_s(BORDURE), bottom=_s(BORDURE))
BORD_BOITE = Border(left=_s(ACIER), right=_s(ACIER),
                    top=_s(ACIER), bottom=_s(ACIER))
BORD_HAUT  = Border(top=_s(OR, 'medium'))
SANS       = Border()


def police(size=10, bold=False, color=ENCRE, italic=False):
    return Font(name=FONT, size=size, bold=bold, color=color, italic=italic)


def fond(color):
    return PatternFill('solid', fgColor=color)


CENTRE   = Alignment(horizontal='center', vertical='center', wrap_text=True)
GAUCHE   = Alignment(horizontal='left',   vertical='center')
DROITE   = Alignment(horizontal='right',  vertical='center')
GAUCHE_W = Alignment(horizontal='left',   vertical='center', wrap_text=True)
HAUT_G   = Alignment(horizontal='left',   vertical='top',    wrap_text=True)


def indent(pos='left', ind=1, wrap=False):
    return Alignment(horizontal=pos, vertical='center', indent=ind, wrap_text=wrap)


# --- Briques de mise en page -------------------------------------------
def bandeau(ws, ligne, col_debut, col_fin, titre, sous_titre='', hauteur=38):
    """Bandeau de titre pleine largeur en haut d'une feuille."""
    for col in range(col_debut, col_fin + 1):
        ws.cell(ligne, col).fill = fond(MARINE)
        ws.cell(ligne + 1, col).fill = fond(ACIER)
    ws.merge_cells(start_row=ligne, start_column=col_debut,
                   end_row=ligne, end_column=col_fin)
    c = ws.cell(ligne, col_debut, titre)
    c.font = police(16, True, BLANC)
    c.alignment = indent('left', 1)
    ws.row_dimensions[ligne].height = hauteur

    ws.merge_cells(start_row=ligne + 1, start_column=col_debut,
                   end_row=ligne + 1, end_column=col_fin)
    c2 = ws.cell(ligne + 1, col_debut, sous_titre)
    c2.font = police(9, False, BLANC)
    c2.alignment = indent('left', 1)
    ws.row_dimensions[ligne + 1].height = 18


def titre_section(ws, ligne, col_debut, col_fin, texte, couleur=ACIER):
    for col in range(col_debut, col_fin + 1):
        ws.cell(ligne, col).fill = fond(couleur)
    ws.merge_cells(start_row=ligne, start_column=col_debut,
                   end_row=ligne, end_column=col_fin)
    c = ws.cell(ligne, col_debut, texte)
    c.font = police(10.5, True, BLANC)
    c.alignment = indent('left', 1)
    ws.row_dimensions[ligne].height = 23


def entetes(ws, ligne, col_debut, libelles, largeurs=None, hauteur=28):
    """Ligne d'en-tetes de tableau."""
    for i, lib in enumerate(libelles):
        col = col_debut + i
        c = ws.cell(ligne, col, lib)
        c.font = police(8.5, True, MARINE)
        c.fill = fond(FOND_2)
        c.alignment = CENTRE
        c.border = Border(left=_s(BLANC), right=_s(BLANC),
                          top=_s(FOND_2), bottom=_s(OR, 'medium'))
        if largeurs:
            ws.column_dimensions[get_column_letter(col)].width = largeurs[i]
    ws.row_dimensions[ligne].height = hauteur


def ligne_total(ws, ligne, col_debut, col_fin, hauteur=23, couleur=MARINE):
    for col in range(col_debut, col_fin + 1):
        c = ws.cell(ligne, col)
        c.fill = fond(couleur)
        c.font = police(10, True, BLANC)
        c.border = Border(top=_s(OR, 'medium'))
    ws.row_dimensions[ligne].height = hauteur


def saisie(cell, fmt=DH):
    """Cellule a remplir par l'utilisateur : fond jaune, texte bleu."""
    cell.fill = fond(JAUNE)
    cell.font = police(10, False, BLEU_SAISIE)
    cell.border = BORD_LEGER
    cell.number_format = fmt
    cell.alignment = DROITE if fmt != TEXTE else indent('left', 1)
    return cell


def calcul(cell, fmt=DH, bold=False, color=ENCRE):
    """Cellule calculee : fond clair, ne pas modifier."""
    cell.fill = fond(BLANC)
    cell.font = police(10, bold, color)
    cell.border = BORD_LEGER
    cell.number_format = fmt
    cell.alignment = DROITE if fmt != TEXTE else indent('left', 1)
    return cell


def libelle(cell, texte=None, bold=False, color=ENCRE, couleur_fond=FOND):
    if texte is not None:
        cell.value = texte
    cell.fill = fond(couleur_fond)
    cell.font = police(9.5, bold, color)
    cell.border = BORD_LEGER
    cell.alignment = indent('left', 1)
    return cell


def note(ws, ligne, col_debut, col_fin, texte, hauteur=28):
    ws.merge_cells(start_row=ligne, start_column=col_debut,
                   end_row=ligne, end_column=col_fin)
    c = ws.cell(ligne, col_debut, texte)
    c.font = police(8.5, False, GRIS, italic=True)
    c.alignment = HAUT_G
    ws.row_dimensions[ligne].height = hauteur
