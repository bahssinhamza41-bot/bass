# -*- coding: utf-8 -*-
"""
Construit le classeur de caisse de PARADISE ALUMINIUM SARL (Casablanca),
pour UN mois.

Structure :
  - une feuille par jour   (onglets 01 a 31)  <- toute la saisie
  - RÉCAP DU JOURNAL       : le mois jour par jour + les reglages
  - TABLEAU DE BORD        : indicateurs et graphiques
  - FOURNISSEURS           : achats, reglements et credit par fournisseur
  - ÉCHÉANCIER             : les cheques et effets, par date d'echeance
  - CHARGES FIXES          : budget par poste contre realise
  - SUIVI MENSUEL          : l'annee mois par mois

Une feuille LISTES, masquee, ne contient que les listes deroulantes.
"""
import os, sys, argparse

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from openpyxl import Workbook
from openpyxl.utils import get_column_letter as GL
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.comments import Comment
from openpyxl.drawing.image import Image as XLImage
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.marker import Marker
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.line import LineProperties
from openpyxl.styles import Alignment

import theme as T
import data_source as D

RACINE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
LOGO_BLANC = os.path.join(RACINE, 'assets', 'logo_paradise_blanc.png')

ap = argparse.ArgumentParser(
    description="Construit le classeur de caisse d'un mois pour Paradise Aluminium.")
ap.add_argument('--annee', type=int, default=2026)
ap.add_argument('--mois', type=int, default=9, choices=range(1, 13), metavar='1-12')
ap.add_argument('--caisse-paradise', type=float, default=None,
                help="Espèces en caisse PARADISE le matin du 1er du mois.")
ap.add_argument('--caisse-zenata', type=float, default=None,
                help="Espèces en caisse ZENATA le matin du 1er du mois.")
ap.add_argument('--autre-caisse', type=float, default=None,
                help="Solde de l'autre caisse (compte courant) au 1er du mois.")
ap.add_argument('--vide', action='store_true',
                help="Ne reprend aucune donnée, même pour septembre 2026.")
ap.add_argument('--sortie', default=None, help="Chemin du fichier à écrire.")
args = ap.parse_args()

ANNEE, MOIS_N = args.annee, args.mois
# La feuille de caisse reprise de PARADISEHAMZA.xlsx est celle du 01/09/2026 :
# elle ne concerne donc que septembre 2026. Le registre des chèques et des
# effets, lui, est un encours : il suit d'un mois sur l'autre.
REPRENDRE = (ANNEE, MOIS_N) == (2026, 9) and not args.vide
REPRENDRE_ECH = not args.vide

def _defaut(valeur, repli):
    return repli if valeur is None else valeur

CAISSE_P = _defaut(args.caisse_paradise,
                   D.CAISSE_PARADISE_OUVERTURE
                   if (ANNEE, MOIS_N) == (2026, 9) else 0)
CAISSE_Z = _defaut(args.caisse_zenata,
                   D.CAISSE_ZENATA_OUVERTURE
                   if (ANNEE, MOIS_N) == (2026, 9) else 0)
AUTRE_C  = _defaut(args.autre_caisse,
                   D.AUTRE_CAISSE_OUVERTURE
                   if (ANNEE, MOIS_N) == (2026, 9) else 0)

MOIS_FR = ['', 'janvier', 'février', 'mars', 'avril', 'mai', 'juin', 'juillet',
           'août', 'septembre', 'octobre', 'novembre', 'décembre']

# Toujours 31 feuilles : un mois plus court neutralise les feuilles en trop.
NB_JOURS = 31
JOURS = [f'{d:02d}' for d in range(1, NB_JOURS + 1)]

# =========================================================================
# Geometrie d'une feuille jour
# =========================================================================
# Bloc achats fournisseurs (colonnes B..J)
L_ACH_1, L_ACH_N, L_ACH_TOT = 8, 27, 28
# Recapitulatif des reglements (colonnes B..D)
L_REG = 30
# Encaissements du jour (colonnes F..J)
L_ENC, L_ENC_H, L_ENC_1, L_ENC_N, L_ENC_TOT = 30, 31, 32, 41, 42
L_OBS, L_OBS_1, L_OBS_N = 37, 38, 44
# Colonne de droite (L..M) : charges fixes, salaires, depenses diverses
L_CF, L_CF_H, L_CF_1 = 6, 7, 8
L_CF_N = L_CF_1 + len(D.CHARGES_FIXES) - 1          # 19
L_CF_TOT = L_CF_N + 1                                # 20
L_SAL, L_SAL_H, L_SAL_1 = 22, 23, 24
L_SAL_N = L_SAL_1 + len(D.PERSONNEL) - 1             # 29
L_SAL_TOT = L_SAL_N + 1                              # 30
L_DIV, L_DIV_H, L_DIV_1, L_DIV_N = 32, 33, 34, 43
L_DIV_TOT = 44
# Bas de page : les caisses
L_CAI, L_CAI_T = 46, 47
L_C1 = 48                                            # premiere ligne des boites
L_SOIR = L_C1 + 8                                    # 56 : solde du soir
L_COMPTE = L_C1 + 9                                  # 57 : especes comptees
L_ECART = L_C1 + 10                                  # 58 : ecart
L_FIN = 60

# Reglages : ils vivent tous sur RÉCAP DU JOURNAL
RCP = "'RÉCAP DU JOURNAL'"
R_AN, R_MOIS = '$E$4', '$E$5'
R_CP, R_CZ, R_AC = '$E$6', '$E$7', '$E$8'
R_JOUR1, R_NBJ = '$J$4', '$J$5'
R_OBJ_ACH, R_OBJ_ENC = '$J$6', '$J$7'

wb = Workbook()
wb.remove(wb.active)


def feuille(nom, couleur, zoom=100):
    ws = wb.create_sheet(nom)
    ws.sheet_properties.tabColor = couleur
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = zoom
    return ws


def larg(ws, mapping):
    for col, w in mapping.items():
        ws.column_dimensions[col].width = w


def somme_jours(colonne, ligne):
    """Additionne la même cellule sur les 31 feuilles jour."""
    lettre = GL(colonne) if isinstance(colonne, int) else colonne
    return '+'.join(f"'{j}'!${lettre}${ligne}" for j in JOURS)


def sumif_jours(col_lib, col_mnt, critere, l1=L_ACH_1, ln=L_ACH_N):
    """Cherche un libellé dans le bloc achats, sur les 31 feuilles jour."""
    a = GL(col_lib) if isinstance(col_lib, int) else col_lib
    b = GL(col_mnt) if isinstance(col_mnt, int) else col_mnt
    return '+'.join(
        f"SUMIF('{j}'!${a}${l1}:${a}${ln},{critere},'{j}'!${b}${l1}:${b}${ln})"
        for j in JOURS)


def zone_impression(ws, ref, paysage=True, une_page=False):
    ws.print_area = ref
    ws.page_setup.orientation = 'landscape' if paysage else 'portrait'
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1 if une_page else 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_options.horizontalCentered = True


# =========================================================================
# 1. LISTES  -- feuille masquée, uniquement les listes déroulantes
# =========================================================================
li = feuille('LISTES', T.GRIS_CLAIR)
larg(li, {'A': 26, 'B': 26, 'C': 22, 'D': 22, 'E': 26, 'F': 60})
COLS_LISTE = [
    ('FOURNISSEURS', D.FOURNISSEURS),
    ('CHARGES FIXES', [c for c, _ in D.CHARGES_FIXES]),
    ('PERSONNEL', [n for n, _ in D.PERSONNEL]),
    ('MODES DE RÈGLEMENT', ['Espèce', 'Chèque', 'Effet', 'Virement', 'Prélèvement']),
    ('ORIGINE DES ENCAISSEMENTS', ['Client comptoir', 'Client sur facture',
                                   'Acompte client', 'Apport associé',
                                   'Retrait banque', 'Autre encaissement']),
]
for i, (titre, valeurs) in enumerate(COLS_LISTE):
    c = li.cell(1, 1 + i, titre)
    c.font = T.police(10, True, T.BLANC)
    c.fill = T.fond(T.ACIER)
    c.alignment = T.CENTRE
    for k, v in enumerate(valeurs):
        li.cell(2 + k, 1 + i, v).font = T.police(10)
li.cell(1, 6, "Cette feuille ne sert qu'aux listes déroulantes des feuilles jour.")
li.cell(2, 6, "Pour ajouter un fournisseur : écrivez-le en bas de la colonne A.")
for r in (1, 2):
    li.cell(r, 6).font = T.police(9, False, T.GRIS, italic=True)

L_FRS = f'LISTES!$A$2:$A${1 + len(D.FOURNISSEURS)}'
L_ORI = f'LISTES!$E$2:$E${1 + len(COLS_LISTE[4][1])}'


# =========================================================================
# 2. FEUILLES JOUR  -- 01 à 31 : toute la saisie du quotidien
# =========================================================================
LARGEURS_JOUR = {'A': 2.6, 'B': 11, 'C': 25, 'D': 14, 'E': 12, 'F': 12,
                 'G': 13, 'H': 12, 'I': 13, 'J': 13, 'K': 2.6,
                 'L': 26, 'M': 14, 'N': 2.6}

CAISSE_P_LIGNES = [
    ('Solde de la veille',                'auto'),
    ('+  Encaissements en espèce',        'auto'),
    ('+  Apports et entrées divers',      'saisie'),
    ('–  Achats réglés en espèce',        'auto'),
    ('–  Charges fixes du jour',          'auto'),
    ('–  Salaires et avances',            'auto'),
    ('–  Dépenses travaux et divers',     'auto'),
    ('–  Versement banque / sorties',     'saisie'),
    ('SOLDE DU SOIR',                     'total'),
    ('Espèces réellement comptées',       'saisie'),
    ('ÉCART  (compté – théorique)',       'ecart'),
]
CAISSE_Z_LIGNES = [
    ('Solde de la veille',                'auto'),
    ('+  Encaissements du jour',          'saisie'),
    ('+  Apports et entrées divers',      'saisie'),
    ('–  Achats réglés en espèce',        'saisie'),
    ('–  Charges et frais du jour',       'saisie'),
    ('–  Salaires et avances',            'saisie'),
    ('–  Dépenses travaux et divers',     'saisie'),
    ('–  Versement banque / sorties',     'saisie'),
    ('SOLDE DU SOIR',                     'total'),
    ('Espèces réellement comptées',       'saisie'),
    ('ÉCART  (compté – théorique)',       'ecart'),
]
RESUME_LIGNES = [
    ('Achats fournisseurs du jour',       'auto'),
    ('dont réglé en espèce',              'auto'),
    ('dont réglé par chèque',             'auto'),
    ('dont réglé par effet',              'auto'),
    ('RESTE DÛ  (crédit fournisseur)',    'alerte'),
    ('Charges fixes',                     'auto'),
    ('Salaires et avances',               'auto'),
    ('Dépenses travaux et divers',        'auto'),
    ('TOTAL DES SORTIES DU JOUR',         'total'),
    ('Encaissements du jour',             'vert'),
    ('CAISSE TOTALE DU SOIR',             'total'),
]


def _boite_titre(ws, ligne, c1, c2, texte, couleur=T.MARINE_2):
    for col in range(c1, c2 + 1):
        ws.cell(ligne, col).fill = T.fond(couleur)
    ws.merge_cells(start_row=ligne, start_column=c1, end_row=ligne, end_column=c2)
    c = ws.cell(ligne, c1, texte)
    c.font = T.police(10, True, T.BLANC)
    c.alignment = T.CENTRE
    ws.row_dimensions[ligne].height = 22


def _boite_ligne(ws, ligne, c_lib1, c_lib2, c_mnt, texte, genre):
    """Une ligne « libellé + montant » dans une boîte de bas de page."""
    if c_lib2 > c_lib1:
        ws.merge_cells(start_row=ligne, start_column=c_lib1,
                       end_row=ligne, end_column=c_lib2)
    lib = ws.cell(ligne, c_lib1, texte)
    mnt = ws.cell(ligne, c_mnt)
    gras = genre == 'total'
    lib.font = T.police(9, gras, T.BLANC if gras else T.ENCRE)
    lib.alignment = T.indent('left', 1)
    for col in range(c_lib1, c_mnt + 1):
        cel = ws.cell(ligne, col)
        cel.border = T.BORD_LEGER
        cel.fill = T.fond(T.MARINE_2 if gras else T.FOND)
    if genre == 'saisie':
        T.saisie(mnt)
    else:
        T.calcul(mnt, T.DH_TOT if gras else T.DH, bold=gras,
                 color=T.BLANC if gras else T.ENCRE)
        if gras:
            mnt.fill = T.fond(T.MARINE_2)
    ws.row_dimensions[ligne].height = 17
    return lib, mnt


def construire_jour(num):
    """Une feuille de caisse pour le jour `num` du mois."""
    ws = feuille(f'{num:02d}', T.OR, zoom=85)
    larg(ws, LARGEURS_JOUR)
    ws.freeze_panes = 'A8'

    veille = f"'{num - 1:02d}'" if num > 1 else None
    date_f = (f'=IF({num}>{RCP}!{R_NBJ},'
              f'"— ce jour n\'existe pas dans le mois —",'
              f'DATE({RCP}!{R_AN},{RCP}!{R_MOIS},{num}))')

    # --- Bandeau ---------------------------------------------------------
    for col in range(1, 15):
        ws.cell(1, col).fill = T.fond(T.MARINE)
        ws.cell(2, col).fill = T.fond(T.ACIER)
    ws.merge_cells('A1:G1')
    c = ws['A1']; c.value = 'JOURNAL DE CAISSE DU JOUR'
    c.font = T.police(15, True, T.BLANC); c.alignment = T.indent('left', 1)
    ws.merge_cells('H1:N1')
    c = ws['H1']; c.value = 'PARADISE ALUMINIUM SARL   ·   Casablanca'
    c.font = T.police(11, True, T.OR); c.alignment = T.indent('right', 1)
    ws.row_dimensions[1].height = 34

    ws.merge_cells('A2:G2')
    c = ws['A2']; c.value = date_f
    c.font = T.police(12, True, T.BLANC); c.number_format = T.JOUR_LONG
    c.alignment = T.indent('left', 1)
    ws.merge_cells('H2:N2')
    c = ws['H2']; c.value = f'Feuille {num:02d}  ·  aluminium, accessoires et vitrage'
    c.font = T.police(9.5, False, T.BLANC); c.alignment = T.indent('right', 1)
    ws.row_dimensions[2].height = 21

    ws.merge_cells('A4:N4')
    c = ws['A4']
    c.value = ("À REMPLIR  ▸  les achats fournisseurs du jour, les charges, les salaires, "
               "les dépenses, les encaissements, puis les espèces comptées le soir dans "
               "chaque caisse.    Cases jaunes = à vous ; cases blanches = calculées.")
    c.font = T.police(9, True, T.MARINE); c.fill = T.fond(T.AMBRE_CLAIR)
    c.alignment = T.indent('left', 1)
    c.border = T.BORD_BOITE
    for col in range(1, 15):
        ws.cell(4, col).fill = T.fond(T.AMBRE_CLAIR)
    ws.row_dimensions[4].height = 26

    # --- Bloc 1 : achats fournisseurs ------------------------------------
    T.titre_section(ws, L_ACH_1 - 2, 2, 10, 'ACHATS FOURNISSEURS DU JOUR')
    T.entetes(ws, L_ACH_1 - 1, 2,
              ['BL N°', 'FOURNISSEUR', 'NET À PAYER', 'ESPÈCE', 'CHÈQUE',
               'N° DU CHÈQUE', 'EFFET', 'N° DE L\'EFFET', 'RESTE À PAYER'])
    for r in range(L_ACH_1, L_ACH_N + 1):
        T.saisie(ws.cell(r, 2), T.TEXTE)
        T.saisie(ws.cell(r, 3), T.TEXTE)
        for col in (4, 5, 6, 8):
            T.saisie(ws.cell(r, col))
        for col in (7, 9):
            T.saisie(ws.cell(r, col), T.TEXTE)
        cel = T.calcul(ws.cell(r, 10))
        cel.value = f'=IF(C{r}="","",D{r}-E{r}-F{r}-H{r})'
        ws.row_dimensions[r].height = 16

    T.ligne_total(ws, L_ACH_TOT, 2, 10)
    ws.merge_cells(start_row=L_ACH_TOT, start_column=2, end_row=L_ACH_TOT, end_column=3)
    c = ws.cell(L_ACH_TOT, 2, 'TOTAL DES ACHATS DU JOUR')
    c.font = T.police(10, True, T.BLANC); c.alignment = T.indent('left', 1)
    for col in (4, 5, 6, 8, 10):
        cel = ws.cell(L_ACH_TOT, col,
                      f'=SUM({GL(col)}{L_ACH_1}:{GL(col)}{L_ACH_N})')
        cel.font = T.police(10, True, T.BLANC)
        cel.number_format = T.DH_TOT
        cel.fill = T.fond(T.MARINE)
        cel.alignment = T.DROITE

    # --- Récapitulatif des règlements (B..D) ------------------------------
    T.titre_section(ws, L_REG, 2, 4, 'RÈGLEMENT DES ACHATS', T.MARINE_2)
    REG = [('Réglé en espèce', f'=E{L_ACH_TOT}'),
           ('Réglé par chèque', f'=F{L_ACH_TOT}'),
           ('Réglé par effet', f'=H{L_ACH_TOT}'),
           ('TOTAL RÉGLÉ', f'=E{L_ACH_TOT}+F{L_ACH_TOT}+H{L_ACH_TOT}'),
           ('RESTE DÛ  (crédit fournisseur)', f'=J{L_ACH_TOT}')]
    for i, (lib, f) in enumerate(REG):
        r = L_REG + 1 + i
        gras = lib.startswith(('TOTAL', 'RESTE'))
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
        T.libelle(ws.cell(r, 2), lib, bold=gras,
                  color=T.ROUGE if lib.startswith('RESTE') else T.ENCRE)
        ws.cell(r, 3).border = T.BORD_LEGER
        ws.cell(r, 3).fill = T.fond(T.FOND)
        cel = T.calcul(ws.cell(r, 4), bold=gras,
                       color=T.ROUGE if lib.startswith('RESTE') else T.ENCRE)
        cel.value = f
        ws.row_dimensions[r].height = 17

    # --- Observations du jour (B..D) --------------------------------------
    T.titre_section(ws, L_OBS, 2, 4, 'OBSERVATIONS DU JOUR', T.GRIS)
    ws.merge_cells(start_row=L_OBS_1, start_column=2, end_row=L_OBS_N, end_column=4)
    for r in range(L_OBS_1, L_OBS_N + 1):
        for col in range(2, 5):
            ws.cell(r, col).fill = T.fond(T.JAUNE)
            ws.cell(r, col).border = T.BORD_LEGER
        ws.row_dimensions[r].height = 15
    c = ws.cell(L_OBS_1, 2)
    c.font = T.police(9.5, False, T.BLEU_SAISIE)
    c.alignment = T.HAUT_G

    # --- Encaissements du jour (F..J) -------------------------------------
    T.titre_section(ws, L_ENC, 6, 10, 'ENCAISSEMENTS DU JOUR', T.VERT)
    T.entetes(ws, L_ENC_H, 6,
              ['CLIENT / ORIGINE', '', 'ESPÈCE', 'CHÈQUE / VIREMENT', 'TOTAL'])
    ws.merge_cells(start_row=L_ENC_H, start_column=6, end_row=L_ENC_H, end_column=7)
    for r in range(L_ENC_1, L_ENC_N + 1):
        ws.merge_cells(start_row=r, start_column=6, end_row=r, end_column=7)
        T.saisie(ws.cell(r, 6), T.TEXTE)
        ws.cell(r, 7).border = T.BORD_LEGER
        ws.cell(r, 7).fill = T.fond(T.JAUNE)
        T.saisie(ws.cell(r, 8))
        T.saisie(ws.cell(r, 9))
        cel = T.calcul(ws.cell(r, 10))
        cel.value = f'=IF(F{r}="","",H{r}+I{r})'
        ws.row_dimensions[r].height = 16
    T.ligne_total(ws, L_ENC_TOT, 6, 10)
    ws.merge_cells(start_row=L_ENC_TOT, start_column=6, end_row=L_ENC_TOT, end_column=7)
    c = ws.cell(L_ENC_TOT, 6, 'TOTAL ENCAISSÉ')
    c.font = T.police(10, True, T.BLANC); c.alignment = T.indent('left', 1)
    for col in (8, 9, 10):
        cel = ws.cell(L_ENC_TOT, col,
                      f'=SUM({GL(col)}{L_ENC_1}:{GL(col)}{L_ENC_N})')
        cel.font = T.police(10, True, T.BLANC)
        cel.number_format = T.DH_TOT
        cel.fill = T.fond(T.MARINE)
        cel.alignment = T.DROITE

    # --- Colonne de droite : charges fixes, salaires, dépenses ------------
    def bloc_droite(ligne_titre, titre, entete, lignes, ligne_total_,
                    libelles=None, couleur=T.ACIER):
        T.titre_section(ws, ligne_titre, 12, 13, titre, couleur)
        T.entetes(ws, ligne_titre + 1, 12, entete, hauteur=22)
        for i, r in enumerate(lignes):
            if libelles is not None:
                T.libelle(ws.cell(r, 12), libelles[i],
                          couleur_fond=T.FOND if i % 2 == 0 else T.BLANC)
            else:
                T.saisie(ws.cell(r, 12), T.TEXTE)
            T.saisie(ws.cell(r, 13))
            ws.row_dimensions[r].height = 16
        T.ligne_total(ws, ligne_total_, 12, 13, hauteur=20)
        c = ws.cell(ligne_total_, 12, 'TOTAL')
        c.font = T.police(9.5, True, T.BLANC); c.alignment = T.indent('left', 1)
        cel = ws.cell(ligne_total_, 13,
                      f'=SUM(M{lignes[0]}:M{lignes[-1]})')
        cel.font = T.police(10, True, T.BLANC)
        cel.number_format = T.DH_TOT
        cel.fill = T.fond(T.MARINE)
        cel.alignment = T.DROITE

    bloc_droite(L_CF, 'CHARGES FIXES DU JOUR', ['POSTE DE CHARGE', 'MONTANT'],
                list(range(L_CF_1, L_CF_N + 1)), L_CF_TOT,
                [lib for lib, _ in D.CHARGES_FIXES])
    bloc_droite(L_SAL, 'SALAIRES ET AVANCES', ['PERSONNE', 'MONTANT'],
                list(range(L_SAL_1, L_SAL_N + 1)), L_SAL_TOT,
                [nom for nom, _ in D.PERSONNEL])
    bloc_droite(L_DIV, 'DÉPENSES TRAVAUX ET DIVERS', ['DÉSIGNATION', 'MONTANT'],
                list(range(L_DIV_1, L_DIV_N + 1)), L_DIV_TOT, None)

    # --- Bas de page : les caisses ----------------------------------------
    T.titre_section(ws, L_CAI, 2, 13, 'LES CAISSES ET LE RÉSUMÉ DU SOIR', T.MARINE)
    _boite_titre(ws, L_CAI_T, 2, 4, 'CAISSE PARADISE')
    _boite_titre(ws, L_CAI_T, 5, 7, 'CAISSE ZENATA')
    _boite_titre(ws, L_CAI_T, 8, 10, 'CONTRÔLE ET AUTRE CAISSE')
    _boite_titre(ws, L_CAI_T, 12, 13, 'RÉSUMÉ DU JOUR', T.MARINE)

    # Caisse PARADISE (B..D)
    for i, (lib, genre) in enumerate(CAISSE_P_LIGNES):
        r = L_C1 + i
        _boite_ligne(ws, r, 2, 3, 4, lib, genre)
    ws.cell(L_C1, 4).value = (f"='{num - 1:02d}'!$D${L_SOIR}" if veille
                              else f'={RCP}!{R_CP}')
    ws.cell(L_C1 + 1, 4).value = f'=H{L_ENC_TOT}'
    ws.cell(L_C1 + 3, 4).value = f'=E{L_ACH_TOT}'
    ws.cell(L_C1 + 4, 4).value = f'=M{L_CF_TOT}'
    ws.cell(L_C1 + 5, 4).value = f'=M{L_SAL_TOT}'
    ws.cell(L_C1 + 6, 4).value = f'=M{L_DIV_TOT}'
    ws.cell(L_SOIR, 4).value = (f'=D{L_C1}+D{L_C1+1}+D{L_C1+2}'
                                f'-D{L_C1+3}-D{L_C1+4}-D{L_C1+5}-D{L_C1+6}-D{L_C1+7}')
    ws.cell(L_ECART, 4).value = f'=IF(D{L_COMPTE}=0,0,D{L_COMPTE}-D{L_SOIR})'
    T.calcul(ws.cell(L_ECART, 4), T.DH, bold=True, color=T.ROUGE)
    ws.cell(L_ECART, 4).fill = T.fond(T.FOND)

    # Caisse ZENATA (E..G)
    for i, (lib, genre) in enumerate(CAISSE_Z_LIGNES):
        r = L_C1 + i
        _boite_ligne(ws, r, 5, 6, 7, lib, genre)
    ws.cell(L_C1, 7).value = (f"='{num - 1:02d}'!$G${L_SOIR}" if veille
                              else f'={RCP}!{R_CZ}')
    ws.cell(L_SOIR, 7).value = (f'=G{L_C1}+G{L_C1+1}+G{L_C1+2}'
                                f'-G{L_C1+3}-G{L_C1+4}-G{L_C1+5}-G{L_C1+6}-G{L_C1+7}')
    ws.cell(L_ECART, 7).value = f'=IF(G{L_COMPTE}=0,0,G{L_COMPTE}-G{L_SOIR})'
    T.calcul(ws.cell(L_ECART, 7), T.DH, bold=True, color=T.ROUGE)
    ws.cell(L_ECART, 7).fill = T.fond(T.FOND)

    # Contrôle et autre caisse (H..J)
    CTRL = [
        ('Caisse PARADISE du soir',        'auto',   f'=D{L_SOIR}'),
        ('Caisse ZENATA du soir',          'auto',   f'=G{L_SOIR}'),
        ('CAISSE TOTALE DU SOIR',          'total',  f'=D{L_SOIR}+G{L_SOIR}'),
        ('Écart total constaté',           'auto',   f'=D{L_ECART}+G{L_ECART}'),
        ('', 'vide', None),
        ('AUTRE CAISSE — solde précédent', 'auto',
         (f"='{num - 1:02d}'!$J${L_C1+8}" if veille else f'={RCP}!{R_AC}')),
        ('+  Entrées du jour',             'saisie', None),
        ('–  Sorties du jour',             'saisie', None),
        ('NOUVEAU SOLDE',                  'total',
         f'=J{L_C1+5}+J{L_C1+6}-J{L_C1+7}'),
    ]
    for i, (lib, genre, f) in enumerate(CTRL):
        r = L_C1 + i
        if genre == 'vide':
            for col in range(8, 11):
                ws.cell(r, col).fill = T.fond(T.BLANC)
            ws.row_dimensions[r].height = 8
            continue
        _boite_ligne(ws, r, 8, 9, 10, lib, genre)
        if f:
            ws.cell(r, 10).value = f
    for r in (L_C1 + 9, L_C1 + 10):
        for col in range(8, 11):
            ws.cell(r, col).fill = T.fond(T.BLANC)

    # Résumé du jour (L..M)
    RES_F = [
        f'=D{L_ACH_TOT}', f'=E{L_ACH_TOT}', f'=F{L_ACH_TOT}', f'=H{L_ACH_TOT}',
        f'=J{L_ACH_TOT}', f'=M{L_CF_TOT}', f'=M{L_SAL_TOT}', f'=M{L_DIV_TOT}',
        (f'=M{L_C1+1}+M{L_C1+2}+M{L_C1+3}+M{L_C1+5}+M{L_C1+6}+M{L_C1+7}'),
        f'=J{L_ENC_TOT}', f'=D{L_SOIR}+G{L_SOIR}',
    ]
    for i, (lib, genre) in enumerate(RESUME_LIGNES):
        r = L_C1 + i
        _boite_ligne(ws, r, 12, 12, 13, lib, genre)
        ws.cell(r, 13).value = RES_F[i]
        if genre == 'alerte':
            T.calcul(ws.cell(r, 13), T.DH, bold=True, color=T.ROUGE)
            ws.cell(r, 13).fill = T.fond(T.FOND)
            ws.cell(r, 12).font = T.police(9, True, T.ROUGE)
        if genre == 'vert':
            T.calcul(ws.cell(r, 13), T.DH, bold=True, color=T.VERT)
            ws.cell(r, 13).fill = T.fond(T.FOND)

    # --- Mises en forme conditionnelles -----------------------------------
    ws.conditional_formatting.add(
        f'J{L_ACH_1}:J{L_ACH_N}',
        CellIsRule(operator='greaterThan', formula=['0'],
                   font=T.police(10, True, T.ROUGE),
                   fill=T.fond(T.ROUGE_CLAIR)))
    for ref in (f'D{L_ECART}', f'G{L_ECART}'):
        ws.conditional_formatting.add(
            ref, CellIsRule(operator='notEqual', formula=['0'],
                            fill=T.fond(T.ROUGE_CLAIR)))
    for ref in (f'D{L_SOIR}', f'G{L_SOIR}', f'J{L_C1+2}'):
        ws.conditional_formatting.add(
            ref, CellIsRule(operator='lessThan', formula=['0'],
                            fill=T.fond(T.ROUGE)))

    # --- Listes déroulantes ------------------------------------------------
    dv = DataValidation(type='list', formula1=f'={L_FRS}', allow_blank=True)
    dv.error = "Choisissez un fournisseur de la liste, ou ajoutez-le dans LISTES."
    dv.errorTitle = 'Fournisseur inconnu'
    dv.showErrorMessage = False
    ws.add_data_validation(dv)
    dv.add(f'C{L_ACH_1}:C{L_ACH_N}')

    dv2 = DataValidation(type='list', formula1=f'={L_ORI}', allow_blank=True)
    dv2.showErrorMessage = False
    ws.add_data_validation(dv2)
    dv2.add(f'F{L_ENC_1}:F{L_ENC_N}')

    ws.cell(L_COMPTE, 4).comment = Comment(
        "Comptez les espèces dans la caisse PARADISE le soir et notez le "
        "montant ici. L'écart se calcule tout seul.", 'Paradise Aluminium', 260, 90)
    zone_impression(ws, f'A1:N{L_FIN}', une_page=True)
    return ws


for _j in range(1, NB_JOURS + 1):
    construire_jour(_j)


# =========================================================================
# 3. RÉCAP DU JOURNAL  -- le mois jour par jour + les réglages
# =========================================================================
rc = feuille('RÉCAP DU JOURNAL', T.MARINE, zoom=90)
larg(rc, {'A': 2.6, 'B': 13, 'C': 11, 'D': 15, 'E': 11, 'F': 11, 'G': 11,
          'H': 13, 'I': 13, 'J': 11, 'K': 13, 'L': 13, 'M': 14, 'N': 14,
          'O': 13, 'P': 14, 'Q': 11, 'R': 2.6})
T.bandeau(rc, 1, 1, 18, 'RÉCAP DU JOURNAL',
          'PARADISE ALUMINIUM SARL  ·  Casablanca  ·  le mois entier, '
          'jour par jour — tout vient des feuilles 01 à 31')

T.titre_section(rc, 3, 2, 17,
                'RÉGLAGES DU MOIS  —  les seules cases à remplir sur cette feuille')
REGLAGES_G = [('Année', ANNEE, T.ENT), ('Mois  (1 à 12)', MOIS_N, T.ENT),
              ('Caisse PARADISE au 1er du mois', CAISSE_P, T.DH),
              ('Caisse ZENATA au 1er du mois', CAISSE_Z, T.DH),
              ("Autre caisse au 1er du mois", AUTRE_C, T.DH)]
for i, (lib, val, fmt) in enumerate(REGLAGES_G):
    r = 4 + i
    rc.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
    T.libelle(rc.cell(r, 2), lib, bold=True)
    for col in (3, 4):
        rc.cell(r, col).border = T.BORD_LEGER
        rc.cell(r, col).fill = T.fond(T.FOND)
    T.saisie(rc.cell(r, 5), fmt).value = val
    rc.row_dimensions[r].height = 19

REGLAGES_D = [
    ('Premier jour du mois', '=DATE($E$4,$E$5,1)', T.DATE),
    ('Nombre de jours du mois', '=DAY(EOMONTH($J$4,0))', T.ENT),
    ("Objectif d'achats du mois", 0, T.DH),
    ("Objectif d'encaissements du mois", 0, T.DH),
    ('Le mois, en clair', '=$J$4', T.MOIS),
]
for i, (lib, val, fmt) in enumerate(REGLAGES_D):
    r = 4 + i
    rc.merge_cells(start_row=r, start_column=7, end_row=r, end_column=9)
    T.libelle(rc.cell(r, 7), lib, bold=True)
    for col in (8, 9):
        rc.cell(r, col).border = T.BORD_LEGER
        rc.cell(r, col).fill = T.fond(T.FOND)
    if i == 4:
        rc.merge_cells(start_row=r, start_column=10, end_row=r, end_column=11)
        rc.cell(r, 11).border = T.BORD_LEGER
        rc.cell(r, 11).fill = T.fond(T.FOND)
    cel = (T.saisie(rc.cell(r, 10), fmt) if isinstance(val, (int, float))
           else T.calcul(rc.cell(r, 10), fmt, bold=True, color=T.MARINE))
    cel.value = val
    cel.alignment = T.CENTRE if fmt in (T.DATE, T.MOIS) else T.DROITE

T.note(rc, 4, 12, 17,
       "Changez l'année, le mois et les trois soldes d'ouverture : tout le classeur "
       "suit — dates, jours de la semaine, feuilles neutralisées pour un mois de "
       "moins de 31 jours.", 30)
T.note(rc, 6, 12, 17,
       "Le report de caisse est le seul lien entre deux mois : le solde du soir du "
       "dernier jour devient le solde d'ouverture du mois suivant.", 30)

ENTETES_RC = ['DATE', 'JOUR', 'ACHATS FOURNIS-\nSEURS', 'DONT\nESPÈCE',
              'DONT\nCHÈQUE', 'DONT\nEFFET', 'RESTE DÛ\nFOURNISSEURS',
              'CHARGES\nFIXES', 'SALAIRES', 'DÉPENSES\n& TRAVAUX',
              'TOTAL DES\nSORTIES', 'ENCAISSE-\nMENTS', 'CAISSE\nPARADISE',
              'CAISSE\nZENATA', 'CAISSE\nTOTALE', 'ÉCART']
L_RC_1, L_RC_N = 12, 42
T.titre_section(rc, 10, 2, 17, 'LE MOIS JOUR PAR JOUR')
T.entetes(rc, 11, 2, ENTETES_RC, hauteur=36)
rc.freeze_panes = 'C12'

# cellule du jour -> (colonne du récap, cellule de la feuille jour)
SOURCES = [
    (4,  f'D{L_ACH_TOT}'), (5,  f'E{L_ACH_TOT}'), (6,  f'F{L_ACH_TOT}'),
    (7,  f'H{L_ACH_TOT}'), (8,  f'J{L_ACH_TOT}'), (9,  f'M{L_CF_TOT}'),
    (10, f'M{L_SAL_TOT}'), (11, f'M{L_DIV_TOT}'), (12, f'M{L_C1+8}'),
    (13, f'J{L_ENC_TOT}'), (14, f'D{L_SOIR}'),    (15, f'G{L_SOIR}'),
    (16, f'J{L_C1+2}'),    (17, f'J{L_C1+3}'),
]
for d in range(1, NB_JOURS + 1):
    r = L_RC_1 + d - 1
    cel = T.calcul(rc.cell(r, 2), T.DATE)
    cel.value = (f'=IF({d}>{R_NBJ},"",DATE({R_AN},{R_MOIS},{d}))')
    cel.alignment = T.CENTRE
    cel = T.calcul(rc.cell(r, 3), T.JOUR_CRT)
    cel.value = f'=IF($B{r}="","",$B{r})'
    cel.alignment = T.CENTRE
    for col, src in SOURCES:
        cel = T.calcul(rc.cell(r, col))
        cel.value = f"=IF($B{r}=\"\",\"\",'{d:02d}'!{src})"
    rc.cell(r, 16).font = T.police(10, True, T.MARINE)
    if d % 2 == 0:
        for col in range(2, 18):
            rc.cell(r, col).fill = T.fond(T.FOND)
    rc.row_dimensions[r].height = 16

L_RC_TOT = L_RC_N + 1
T.ligne_total(rc, L_RC_TOT, 2, 17, hauteur=24)
rc.merge_cells(start_row=L_RC_TOT, start_column=2, end_row=L_RC_TOT, end_column=3)
c = rc.cell(L_RC_TOT, 2, 'TOTAL DU MOIS')
c.font = T.police(10, True, T.BLANC); c.alignment = T.indent('left', 1)
for col in list(range(4, 14)) + [17]:
    cel = rc.cell(L_RC_TOT, col, f'=SUM({GL(col)}{L_RC_1}:{GL(col)}{L_RC_N})')
    cel.font = T.police(10, True, T.BLANC)
    cel.number_format = T.DH_TOT
    cel.fill = T.fond(T.MARINE)
    cel.alignment = T.DROITE
for col in (14, 15, 16):
    cel = rc.cell(L_RC_TOT, col,
                  f'=INDEX({GL(col)}{L_RC_1}:{GL(col)}{L_RC_N},{R_NBJ})')
    cel.font = T.police(10, True, T.OR)
    cel.number_format = T.DH_TOT
    cel.fill = T.fond(T.MARINE)
    cel.alignment = T.DROITE
T.note(rc, L_RC_TOT + 1, 2, 17,
       "Les trois colonnes de caisse ne s'additionnent pas : la ligne de total "
       "reprend le solde du dernier jour du mois, c'est-à-dire ce qui reste "
       "réellement en caisse.", 26)

rc.conditional_formatting.add(
    f'H{L_RC_1}:H{L_RC_N}',
    CellIsRule(operator='greaterThan', formula=['0'],
               fill=T.fond(T.ROUGE_CLAIR), font=T.police(10, True, T.ROUGE)))
rc.conditional_formatting.add(
    f'M{L_RC_1}:M{L_RC_N}',
    CellIsRule(operator='greaterThan', formula=['0'],
               font=T.police(10, True, T.VERT)))
rc.conditional_formatting.add(
    f'Q{L_RC_1}:Q{L_RC_N}',
    FormulaRule(formula=[f'AND($B{L_RC_1}<>"",$Q{L_RC_1}<>0)'],
                fill=T.fond(T.ROUGE_CLAIR), font=T.police(10, True, T.ROUGE)))
for ref in (f'N{L_RC_1}:N{L_RC_N}', f'O{L_RC_1}:O{L_RC_N}', f'P{L_RC_1}:P{L_RC_N}'):
    rc.conditional_formatting.add(
        ref, CellIsRule(operator='lessThan', formula=['0'],
                        fill=T.fond(T.ROUGE), font=T.police(10, True, T.BLANC)))
# Colonnes de service, masquées : elles renvoient #N/A pour les jours qui
# n'existent pas dans le mois, ce qui évite aux graphiques de retomber à zéro
# à la fin d'un mois de 28 ou 30 jours.
SERIES = [(19, 'Achats fournisseurs', 'D'), (20, 'Caisse totale', 'P'),
          (21, 'Encaissements', 'M'), (22, 'Sorties', 'L')]
for col, nom, src in SERIES:
    rc.cell(11, col, nom)
    rc.column_dimensions[GL(col)].hidden = True
    for d in range(1, NB_JOURS + 1):
        r = L_RC_1 + d - 1
        if col == 20 and d > 1:
            # La courbe de caisse reste plate sur les jours qui n'existent pas
            # plutôt que de retomber à zéro en fin de mois court.
            rc.cell(r, col, f'=IF($B{r}="",{GL(col)}{r - 1},${src}{r})')
        else:
            rc.cell(r, col, f'=IF($B{r}="",NA(),${src}{r})')

zone_impression(rc, f'A1:R{L_RC_TOT + 1}')


# =========================================================================
# 4. FOURNISSEURS  -- achats, règlements et crédit, fournisseur par fournisseur
# =========================================================================
fo = feuille('FOURNISSEURS', T.ACIER, zoom=95)
larg(fo, {'A': 2.6, 'B': 30, 'C': 16, 'D': 14, 'E': 14, 'F': 14, 'G': 15,
          'H': 15, 'I': 13, 'J': 2.6})
T.bandeau(fo, 1, 1, 10, 'FOURNISSEURS DU MOIS',
          'Ce que nous leur avons acheté, ce que nous leur avons réglé, '
          'et ce que nous leur devons encore')

L_FO_1 = 6
NOMS_FRS = list(D.FOURNISSEURS)
L_FO_N = L_FO_1 + len(NOMS_FRS) - 1
L_FO_TOT = L_FO_N + 1

T.titre_section(fo, 4, 2, 9, 'ACHATS ET RÈGLEMENTS PAR FOURNISSEUR')
T.entetes(fo, 5, 2,
          ['FOURNISSEUR', 'ACHATS DU MOIS', 'RÉGLÉ EN\nESPÈCE',
           'RÉGLÉ PAR\nCHÈQUE', 'RÉGLÉ PAR\nEFFET', 'TOTAL RÉGLÉ',
           'RESTE DÛ', 'PART DES\nACHATS'], hauteur=32)
fo.freeze_panes = 'B6'

for i, nom in enumerate(NOMS_FRS):
    r = L_FO_1 + i
    T.libelle(fo.cell(r, 2), nom, bold=True,
              couleur_fond=T.FOND if i % 2 == 0 else T.BLANC)
    for col, src in ((3, 'D'), (4, 'E'), (5, 'F'), (6, 'H')):
        cel = T.calcul(fo.cell(r, col))
        cel.value = '=' + sumif_jours('C', src, f'$B{r}')
        if i % 2 == 0:
            cel.fill = T.fond(T.FOND)
    cel = T.calcul(fo.cell(r, 7), bold=True)
    cel.value = f'=D{r}+E{r}+F{r}'
    cel = T.calcul(fo.cell(r, 8), bold=True)
    cel.value = f'=C{r}-G{r}'
    cel = T.calcul(fo.cell(r, 9), T.PCT)
    cel.value = f'=IF($C${L_FO_TOT}=0,0,C{r}/$C${L_FO_TOT})'
    if i % 2 == 0:
        for col in (7, 8, 9):
            fo.cell(r, col).fill = T.fond(T.FOND)
    fo.row_dimensions[r].height = 17

T.ligne_total(fo, L_FO_TOT, 2, 9, hauteur=24)
c = fo.cell(L_FO_TOT, 2, 'TOTAL FOURNISSEURS')
c.font = T.police(10, True, T.BLANC); c.alignment = T.indent('left', 1)
for col in range(3, 9):
    cel = fo.cell(L_FO_TOT, col, f'=SUM({GL(col)}{L_FO_1}:{GL(col)}{L_FO_N})')
    cel.font = T.police(10, True, T.BLANC)
    cel.number_format = T.DH_TOT
    cel.fill = T.fond(T.MARINE)
    cel.alignment = T.DROITE
cel = fo.cell(L_FO_TOT, 9, f'=IF($C${L_FO_TOT}=0,0,1)')
cel.font = T.police(10, True, T.BLANC); cel.number_format = T.PCT
cel.fill = T.fond(T.MARINE); cel.alignment = T.DROITE

T.titre_section(fo, L_FO_TOT + 2, 2, 9, 'CONTRÔLE')
CTRL_FO = [
    ("Total des achats saisis sur les 31 feuilles jour",
     '=' + somme_jours('D', L_ACH_TOT)),
    ("Total repris ci-dessus, fournisseur par fournisseur", f'=C{L_FO_TOT}'),
    ("ÉCART  —  doit rester à zéro",
     f'=' + somme_jours('D', L_ACH_TOT) + f'-C{L_FO_TOT}'),
]
for i, (lib, f) in enumerate(CTRL_FO):
    r = L_FO_TOT + 3 + i
    fo.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
    T.libelle(fo.cell(r, 2), lib, bold=(i == 2))
    for col in range(3, 7):
        fo.cell(r, col).border = T.BORD_LEGER
        fo.cell(r, col).fill = T.fond(T.FOND)
    cel = T.calcul(fo.cell(r, 7), bold=(i == 2),
                   color=T.ROUGE if i == 2 else T.ENCRE)
    cel.value = f
    fo.row_dimensions[r].height = 18
T.note(fo, L_FO_TOT + 6, 2, 9,
       "Un écart non nul veut dire qu'un fournisseur a été écrit à la main sur une "
       "feuille jour sans figurer dans la liste ci-dessus. Ajoutez-le dans la feuille "
       "LISTES (colonne A), puis dans ce tableau.", 30)

# Tableau de service, masqué : le classement des dix premiers fournisseurs,
# qui alimente le graphique du tableau de bord. La clé de la colonne J rend
# chaque montant unique pour que le rapprochement par MATCH reste juste.
for i in range(len(NOMS_FRS)):
    r = L_FO_1 + i
    fo.cell(r, 10, f'=$C{r}-ROW()/1000000')
L_TOP_1 = L_FO_1
fo.cell(L_TOP_1 - 1, 12, 'Fournisseur')
fo.cell(L_TOP_1 - 1, 13, 'Achats du mois')
for i in range(10):
    k = 10 - i                       # le plus gros finit en haut du graphique
    r = L_TOP_1 + i
    cle = f'LARGE($J${L_FO_1}:$J${L_FO_N},{k})'
    pos = f'MATCH({cle},$J${L_FO_1}:$J${L_FO_N},0)'
    fo.cell(r, 12, f'=INDEX($B${L_FO_1}:$B${L_FO_N},{pos})')
    fo.cell(r, 13, f'=INDEX($C${L_FO_1}:$C${L_FO_N},{pos})')
L_TOP_N = L_TOP_1 + 9
for col in ('J', 'L', 'M'):
    fo.column_dimensions[col].hidden = True

fo.conditional_formatting.add(
    f'H{L_FO_1}:H{L_FO_N}',
    CellIsRule(operator='greaterThan', formula=['0'],
               fill=T.fond(T.ROUGE_CLAIR), font=T.police(10, True, T.ROUGE)))
fo.conditional_formatting.add(
    f'C{L_FO_1}:C{L_FO_N}',
    CellIsRule(operator='greaterThan', formula=['0'],
               font=T.police(10, True, T.MARINE)))
zone_impression(fo, f'A1:J{L_FO_TOT + 6}')


# =========================================================================
# 5. ÉCHÉANCIER  -- les chèques et les effets, par date d'échéance
# =========================================================================
ec = feuille('ÉCHÉANCIER', T.ROUGE, zoom=95)
larg(ec, {'A': 2.6, 'B': 13, 'C': 22, 'D': 30, 'E': 17, 'F': 15, 'G': 15,
          'H': 18, 'I': 46, 'J': 2.6})
T.bandeau(ec, 1, 1, 10, 'ÉCHÉANCIER DES CHÈQUES ET DES EFFETS',
          "Tout ce que la société doit encore régler, avec sa date d'échéance — "
          "l'état se met à jour tout seul à l'ouverture du classeur")

NB_ECH = max(len(D.ECHEANCIER) + 40, 140) if REPRENDRE_ECH else 140
L_EC_1 = 9
L_EC_N = L_EC_1 + NB_ECH - 1
PL_F = f'$F${L_EC_1}:$F${L_EC_N}'
PL_H = f'$H${L_EC_1}:$H${L_EC_N}'
PL_G = f'$G${L_EC_1}:$G${L_EC_N}'

T.titre_section(ec, 3, 2, 9, "OÙ NOUS EN SOMMES")
CARTES_EC = [
    ('TOTAL RESTANT DÛ', f'=SUM({PL_F})', T.MARINE),
    ('DONT EN RETARD', f'=SUMIF({PL_H},"EN RETARD",{PL_F})', T.ROUGE),
    ('À ÉCHOIR SOUS 30 JOURS', f'=SUMIF({PL_H},"À ÉCHOIR ≤ 30 J",{PL_F})', T.AMBRE),
    ('DÉJÀ PAYÉ (cumul)', f'=SUM({PL_G})', T.VERT),
]
for i, (lib, f, coul) in enumerate(CARTES_EC):
    r = 4
    c1 = 2 + i * 2
    ec.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c1 + 1)
    ec.merge_cells(start_row=r + 1, start_column=c1, end_row=r + 1, end_column=c1 + 1)
    t = ec.cell(r, c1, lib)
    t.font = T.police(8.5, True, T.BLANC); t.fill = T.fond(coul)
    t.alignment = T.CENTRE
    ec.row_dimensions[r].height = 18
    v = ec.cell(r + 1, c1, f)
    v.font = T.police(14, True, coul); v.number_format = T.DH_TOT
    v.fill = T.fond(T.FOND); v.alignment = T.CENTRE
    v.border = T.BORD_LEGER
    ec.row_dimensions[r + 1].height = 26
    for col in (c1, c1 + 1):
        ec.cell(r, col).fill = T.fond(coul)
        ec.cell(r + 1, col).fill = T.fond(T.FOND)
        ec.cell(r + 1, col).border = T.BORD_LEGER

T.entetes(ec, L_EC_1 - 1, 2,
          ["DATE D'ÉCHÉANCE", 'TYPE DE RÈGLEMENT', 'BÉNÉFICIAIRE / OBJET',
           'N° DU CHÈQUE\nOU DE L\'EFFET', 'MONTANT\nRESTANT DÛ',
           'MONTANT\nDÉJÀ PAYÉ', 'ÉTAT', 'OBSERVATIONS'], hauteur=32)
ec.freeze_panes = 'B9'

ETAT = ('=IF(AND($B{r}="",$D{r}=""),"",'
        'IF($F{r}<=0,"PAYÉ",'
        'IF($B{r}="","SANS DATE",'
        'IF($B{r}<TODAY(),"EN RETARD",'
        'IF($B{r}<=TODAY()+30,"À ÉCHOIR ≤ 30 J","À VENIR")))))')

import datetime as _dt
lignes_ech = D.ECHEANCIER if REPRENDRE_ECH else []
for i in range(NB_ECH):
    r = L_EC_1 + i
    T.saisie(ec.cell(r, 2), T.DATE).alignment = T.CENTRE
    T.saisie(ec.cell(r, 3), T.TEXTE)
    T.saisie(ec.cell(r, 4), T.TEXTE)
    T.saisie(ec.cell(r, 5), T.TEXTE).alignment = T.CENTRE
    T.saisie(ec.cell(r, 6))
    T.saisie(ec.cell(r, 7))
    cel = T.calcul(ec.cell(r, 8), T.TEXTE, bold=True)
    cel.value = ETAT.format(r=r)
    cel.alignment = T.CENTRE
    T.saisie(ec.cell(r, 9), T.TEXTE)
    ec.row_dimensions[r].height = 16
    if i < len(lignes_ech):
        typ, ben, num, du, dat, paye, obs = lignes_ech[i]
        if dat:
            ec.cell(r, 2).value = _dt.datetime.strptime(dat, '%Y-%m-%d')
        ec.cell(r, 3).value = typ or ''
        ec.cell(r, 4).value = ben or ''
        ec.cell(r, 5).value = str(num) if num else ''
        ec.cell(r, 6).value = du or 0
        ec.cell(r, 7).value = paye or 0
        ec.cell(r, 9).value = obs or ''

for texte, remplissage, encre in (
        ('EN RETARD', T.ROUGE_CLAIR, T.ROUGE),
        ('À ÉCHOIR ≤ 30 J', T.AMBRE_CLAIR, T.AMBRE),
        ('PAYÉ', T.VERT_CLAIR, T.VERT),
        ('SANS DATE', T.FOND_2, T.GRIS)):
    ec.conditional_formatting.add(
        f'B{L_EC_1}:I{L_EC_N}',
        FormulaRule(formula=[f'$H{L_EC_1}="{texte}"'],
                    fill=T.fond(remplissage), font=T.police(9.5, True, encre),
                    stopIfTrue=False))

ec.auto_filter.ref = f'B{L_EC_1 - 1}:I{L_EC_N}'
T.note(ec, L_EC_N + 2, 2, 9,
       "Le total à venir au-delà de 30 jours se lit en retirant du total restant dû "
       "ce qui est en retard et ce qui échoit sous 30 jours.    "
       "MONTANT RESTANT DÛ = ce qui n'est pas encore débité. Quand le chèque ou "
       "l'effet est payé, portez le montant dans « déjà payé » et ramenez le "
       "restant dû à zéro : la ligne passe en vert toute seule.", 30)
zone_impression(ec, f'A1:J{L_EC_N + 2}')


# =========================================================================
# 6. CHARGES FIXES  -- budget par poste contre réalisé
# =========================================================================
cf = feuille('CHARGES FIXES', T.OR, zoom=95)
larg(cf, {'A': 2.6, 'B': 32, 'C': 16, 'D': 16, 'E': 16, 'F': 14, 'G': 40, 'H': 2.6})
T.bandeau(cf, 1, 1, 8, 'CHARGES FIXES ET SALAIRES',
          'Ce que le mois devait coûter, ce qu\'il a réellement coûté')

ENT_CF = ['POSTE', 'BUDGET DU MOIS', 'RÉALISÉ', 'ÉCART', '% DU BUDGET']


def table_budget(ligne_titre, titre, postes, ligne_source_debut, ligne_tot_jour):
    """Un bloc « budget contre réalisé » adossé aux feuilles jour."""
    T.titre_section(cf, ligne_titre, 2, 6, titre)
    T.entetes(cf, ligne_titre + 1, 2, ENT_CF, hauteur=26)
    r0 = ligne_titre + 2
    for i, (lib, budget) in enumerate(postes):
        r = r0 + i
        T.libelle(cf.cell(r, 2), lib, bold=True,
                  couleur_fond=T.FOND if i % 2 == 0 else T.BLANC)
        T.saisie(cf.cell(r, 3)).value = budget
        cel = T.calcul(cf.cell(r, 4), bold=True)
        cel.value = '=' + somme_jours('M', ligne_source_debut + i)
        T.calcul(cf.cell(r, 5)).value = f'=C{r}-D{r}'
        T.calcul(cf.cell(r, 6), T.PCT).value = f'=IF(C{r}=0,0,D{r}/C{r})'
        if i % 2 == 0:
            for col in (4, 5, 6):
                cf.cell(r, col).fill = T.fond(T.FOND)
        cf.row_dimensions[r].height = 17
    rt = r0 + len(postes)
    T.ligne_total(cf, rt, 2, 6, hauteur=23)
    c = cf.cell(rt, 2, 'TOTAL')
    c.font = T.police(10, True, T.BLANC); c.alignment = T.indent('left', 1)
    for col in (3, 4, 5):
        cel = cf.cell(rt, col, f'=SUM({GL(col)}{r0}:{GL(col)}{rt - 1})')
        cel.font = T.police(10, True, T.BLANC); cel.number_format = T.DH_TOT
        cel.fill = T.fond(T.MARINE); cel.alignment = T.DROITE
    cel = cf.cell(rt, 6, f'=IF(C{rt}=0,0,D{rt}/C{rt})')
    cel.font = T.police(10, True, T.BLANC); cel.number_format = T.PCT
    cel.fill = T.fond(T.MARINE); cel.alignment = T.DROITE
    cf.conditional_formatting.add(
        f'D{r0}:D{rt - 1}',
        FormulaRule(formula=[f'AND($C{r0}>0,$D{r0}>$C{r0})'],
                    fill=T.fond(T.ROUGE_CLAIR), font=T.police(10, True, T.ROUGE)))
    return rt


L_CF_TOT_B = table_budget(4, 'CHARGES FIXES DU MOIS', D.CHARGES_FIXES,
                          L_CF_1, L_CF_TOT)
L_SAL_TOT_B = table_budget(L_CF_TOT_B + 2, 'SALAIRES ET AVANCES', D.PERSONNEL,
                           L_SAL_1, L_SAL_TOT)

L_SYN = L_SAL_TOT_B + 2
T.titre_section(cf, L_SYN, 2, 6, 'LE MOIS EN UN COUP D\'ŒIL', T.MARINE)
SYNTHESE = [
    ('Charges fixes réalisées', f'=D{L_CF_TOT_B}', T.ENCRE),
    ('Salaires et avances', f'=D{L_SAL_TOT_B}', T.ENCRE),
    ('Dépenses travaux et divers', '=' + somme_jours('M', L_DIV_TOT), T.ENCRE),
    ('CHARGES DE STRUCTURE DU MOIS',
     f'=D{L_CF_TOT_B}+D{L_SAL_TOT_B}+' + somme_jours('M', L_DIV_TOT), T.MARINE),
    ('Achats fournisseurs réglés', '=' + somme_jours('E', L_ACH_TOT) + '+'
     + somme_jours('F', L_ACH_TOT) + '+' + somme_jours('H', L_ACH_TOT), T.ENCRE),
    ('TOTAL DES SORTIES DU MOIS', f"={RCP}!$L${L_RC_TOT}", T.MARINE),
    ('Encaissements du mois', f"={RCP}!$M${L_RC_TOT}", T.VERT),
    ('RÉSULTAT DE TRÉSORERIE DU MOIS',
     f"={RCP}!$M${L_RC_TOT}-{RCP}!$L${L_RC_TOT}", T.MARINE),
]
for i, (lib, f, coul) in enumerate(SYNTHESE):
    r = L_SYN + 1 + i
    gras = lib.isupper()
    cf.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
    T.libelle(cf.cell(r, 2), lib, bold=gras, color=coul if gras else T.ENCRE,
              couleur_fond=T.FOND_2 if gras else T.FOND)
    for col in (3, 4, 5):
        cf.cell(r, col).border = T.BORD_LEGER
        cf.cell(r, col).fill = T.fond(T.FOND_2 if gras else T.FOND)
    cel = T.calcul(cf.cell(r, 6), T.DH, bold=gras, color=coul)
    cel.value = f
    cel.fill = T.fond(T.FOND_2 if gras else T.BLANC)
    cf.row_dimensions[r].height = 19
cf.conditional_formatting.add(
    f'F{L_SYN + 8}',
    CellIsRule(operator='lessThan', formula=['0'],
               fill=T.fond(T.ROUGE_CLAIR), font=T.police(10, True, T.ROUGE)))
T.note(cf, L_SYN + 10, 2, 6,
       "Le budget est une colonne à vous : ajustez-le une fois pour toutes, il sert "
       "de repère tous les mois. Un poste dépassé se colore en rouge de lui-même.", 30)
zone_impression(cf, f'A1:H{L_SYN + 10}', paysage=False)


# =========================================================================
# 7. SUIVI MENSUEL  -- l'année mois par mois
# =========================================================================
sm = feuille('SUIVI MENSUEL', T.ACIER, zoom=95)
larg(sm, {'A': 2.6, 'B': 16, 'C': 16, 'D': 15, 'E': 13, 'F': 15, 'G': 16,
          'H': 16, 'I': 16, 'J': 14, 'K': 2.6})
T.bandeau(sm, 1, 1, 11, f'SUIVI MENSUEL  —  {ANNEE}',
          "Un classeur par mois ; cette page rassemble l'année. "
          "Le mois en cours se remplit tout seul.")

T.titre_section(sm, 4, 2, 10, f"L'ANNÉE {ANNEE} MOIS PAR MOIS")
T.entetes(sm, 5, 2,
          ['MOIS', 'ACHATS\nFOURNISSEURS', 'CHARGES\nFIXES', 'SALAIRES',
           'DÉPENSES\n& TRAVAUX', 'TOTAL DES\nSORTIES', 'ENCAISSEMENTS',
           'CAISSE EN FIN\nDE MOIS', 'RÉSULTAT DE\nTRÉSORERIE'], hauteur=34)
L_SM_1 = 6
for i in range(12):
    r = L_SM_1 + i
    T.libelle(sm.cell(r, 2), MOIS_FR[i + 1].capitalize(), bold=True,
              couleur_fond=T.FOND if i % 2 == 0 else T.BLANC)
    for col in range(3, 10):
        if i + 1 == MOIS_N:
            src = {3: f'$D${L_RC_TOT}', 4: f'$I${L_RC_TOT}', 5: f'$J${L_RC_TOT}',
                   6: f'$K${L_RC_TOT}', 7: f'$L${L_RC_TOT}', 8: f'$M${L_RC_TOT}',
                   9: f'$P${L_RC_TOT}'}[col]
            cel = T.calcul(sm.cell(r, col), bold=True, color=T.MARINE)
            cel.value = f'={RCP}!{src}'
            cel.fill = T.fond(T.FOND_2)
        else:
            cel = T.saisie(sm.cell(r, col))
    if i + 1 == MOIS_N:
        T.libelle(sm.cell(r, 2), bold=True, color=T.MARINE, couleur_fond=T.FOND_2)
    cel = T.calcul(sm.cell(r, 10), bold=True)
    cel.value = f'=H{r}-G{r}'
    sm.row_dimensions[r].height = 18
L_SM_TOT = L_SM_1 + 12
T.ligne_total(sm, L_SM_TOT, 2, 10, hauteur=24)
c = sm.cell(L_SM_TOT, 2, f'TOTAL {ANNEE}')
c.font = T.police(10, True, T.BLANC); c.alignment = T.indent('left', 1)
for col in list(range(3, 9)) + [10]:
    cel = sm.cell(L_SM_TOT, col, f'=SUM({GL(col)}{L_SM_1}:{GL(col)}{L_SM_1 + 11})')
    cel.font = T.police(10, True, T.BLANC); cel.number_format = T.DH_TOT
    cel.fill = T.fond(T.MARINE); cel.alignment = T.DROITE
cel = sm.cell(L_SM_TOT, 9,
              f'=IFERROR(LOOKUP(2,1/($I${L_SM_1}:$I${L_SM_1 + 11}<>""),'
              f'$I${L_SM_1}:$I${L_SM_1 + 11}),0)')
cel.font = T.police(10, True, T.OR); cel.number_format = T.DH_TOT
cel.fill = T.fond(T.MARINE); cel.alignment = T.DROITE

L_REF = L_SM_TOT + 2
T.titre_section(sm, L_REF, 2, 10,
                'RAPPEL  —  chiffres du dossier bancaire, janvier à juillet 2026')
REF = [
    ("Chiffre d'affaires facturé sur 7 mois", D.REFERENCE_2026['ca_facture']),
    ("Achats fournisseurs sur 7 mois (TTC)", D.REFERENCE_2026['achats_ttc']),
    ("Versements d'espèces en banque sur 7 mois", D.REFERENCE_2026['versements_especes']),
    ("Flux d'affaires total encaissé sur 7 mois", D.REFERENCE_2026['flux_total']),
    ("Moyenne d'achats par mois",
     round(D.REFERENCE_2026['achats_ttc'] / D.REFERENCE_2026['mois'])),
    ("Clients facturés / fournisseurs référencés", None),
]
for i, (lib, val) in enumerate(REF):
    r = L_REF + 1 + i
    sm.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
    T.libelle(sm.cell(r, 2), lib, bold=(i >= 3))
    for col in (3, 4, 5):
        sm.cell(r, col).border = T.BORD_LEGER
        sm.cell(r, col).fill = T.fond(T.FOND)
    cel = T.calcul(sm.cell(r, 6), T.DH, bold=(i >= 3), color=T.MARINE)
    if val is None:
        cel.number_format = T.TEXTE
        cel.value = (f"{D.REFERENCE_2026['clients']} clients  ·  "
                     f"{D.REFERENCE_2026['fournisseurs']} fournisseurs")
        cel.alignment = T.CENTRE
    else:
        cel.value = val
    sm.row_dimensions[r].height = 18
T.note(sm, L_REF + 8, 2, 10,
       "Source : dossier de présentation de la société (demande de découvert "
       "bancaire, septembre 2026). Ces chiffres servent de repère, ils ne se "
       "recalculent pas.", 30)
zone_impression(sm, f'A1:K{L_REF + 8}')


# =========================================================================
# 8. TABLEAU DE BORD  -- le mois en huit chiffres et quatre graphiques
# =========================================================================
tb = feuille('TABLEAU DE BORD', T.MARINE, zoom=90)
larg(tb, {'A': 2.6, **{GL(c): 12.5 for c in range(2, 14)}, 'N': 2.6})
tb.sheet_view.showGridLines = False

for r, h in ((1, 10), (2, 56), (3, 10)):
    tb.row_dimensions[r].height = h
    for col in range(1, 15):
        tb.cell(r, col).fill = T.fond(T.MARINE)
tb.merge_cells('F2:M2')
c = tb['F2']; c.value = 'TABLEAU DE BORD DU MOIS'
c.font = T.police(19, True, T.BLANC)
c.alignment = T.indent('right', 1)

if os.path.exists(LOGO_BLANC):
    img = XLImage(LOGO_BLANC)
    img.width, img.height = 152, 83
    img.anchor = 'B2'
    tb.add_image(img)
else:                                     # repli si le logo est absent
    tb.merge_cells('B2:E2')
    c = tb['B2']; c.value = 'PARADISE ALUMINIUM'
    c.font = T.police(18, True, T.BLANC); c.alignment = T.indent('left', 1)

tb.merge_cells('B4:M4')
c = tb['B4']
c.value = (f'=UPPER(TEXT({RCP}!{R_JOUR1},"[$-40C]MMMM YYYY"))'
           f'&"     ·     PARADISE ALUMINIUM SARL     ·     '
           f'Négoce et transformation d\'aluminium     ·     Casablanca"')
c.font = T.police(10, True, T.MARINE)
c.alignment = T.CENTRE
for col in range(2, 14):
    tb.cell(4, col).fill = T.fond(T.FOND_2)
tb.row_dimensions[4].height = 24

CARTES = [
    ('ACHATS FOURNISSEURS', f'={RCP}!$D${L_RC_TOT}', 'net à payer, cumul du mois', T.MARINE),
    ('ENCAISSEMENTS', f'={RCP}!$M${L_RC_TOT}', 'clients et apports', T.VERT),
    ('TOTAL DES SORTIES', f'={RCP}!$L${L_RC_TOT}', 'achats réglés + charges', T.ACIER),
    ('RÉSULTAT DE TRÉSORERIE', f'={RCP}!$M${L_RC_TOT}-{RCP}!$L${L_RC_TOT}',
     'encaissements – sorties', T.MARINE),
    ('RESTE DÛ AUX FOURNISSEURS', f'={RCP}!$H${L_RC_TOT}',
     'crédit fournisseur du mois', T.ROUGE),
    ('CHARGES FIXES ET SALAIRES', f'={RCP}!$I${L_RC_TOT}+{RCP}!$J${L_RC_TOT}',
     'le coût de la structure', T.AMBRE),
    ('CAISSE TOTALE EN FIN DE MOIS', f'={RCP}!$P${L_RC_TOT}',
     'PARADISE + ZENATA', T.MARINE),
    ('ÉCART DE CAISSE CUMULÉ', f'={RCP}!$Q${L_RC_TOT}',
     'compté – théorique', T.ROUGE),
]
for i, (titre, f, sous, coul) in enumerate(CARTES):
    bloc, pos = divmod(i, 4)
    c1 = 2 + pos * 3
    c2 = c1 + 2
    r0 = 6 + bloc * 4
    for dr, h in ((0, 17), (1, 30), (2, 15)):
        tb.merge_cells(start_row=r0 + dr, start_column=c1,
                       end_row=r0 + dr, end_column=c2)
        tb.row_dimensions[r0 + dr].height = h
    t = tb.cell(r0, c1, titre)
    t.font = T.police(8.5, True, T.BLANC); t.alignment = T.CENTRE
    v = tb.cell(r0 + 1, c1, f)
    v.font = T.police(17, True, coul); v.number_format = T.DH_TOT
    v.alignment = T.CENTRE
    s = tb.cell(r0 + 2, c1, sous)
    s.font = T.police(8, False, T.GRIS, italic=True); s.alignment = T.CENTRE
    for col in range(c1, c2 + 1):
        tb.cell(r0, col).fill = T.fond(coul)
        tb.cell(r0 + 1, col).fill = T.fond(T.BLANC)
        tb.cell(r0 + 2, col).fill = T.fond(T.BLANC)
        tb.cell(r0 + 1, col).border = T.BORD_LEGER
        tb.cell(r0 + 2, col).border = T.BORD_LEGER
    tb.row_dimensions[r0 + 3].height = 6

tb.conditional_formatting.add(
    'B7:M11', CellIsRule(operator='lessThan', formula=['0'],
                         font=T.police(17, True, T.ROUGE)))


def _gp(couleur):
    return GraphicalProperties(solidFill=couleur)


def habiller(ch, titre, largeur=16.4, hauteur=8.4, legende=None):
    ch.title = titre
    ch.style = None
    ch.width, ch.height = largeur, hauteur
    # Les séries vivent dans des colonnes masquées : sans cela, ni Excel ni
    # LibreOffice ne les tracent.
    ch.visible_cells_only = False      # les séries vivent en colonnes masquées
    ch.display_blanks = 'gap'
    if legende is None:
        ch.legend = None
    else:
        ch.legend.position = legende
        ch.legend.overlay = False
    ch.y_axis.majorGridlines.spPr = GraphicalProperties(
        ln=LineProperties(solidFill=T.BORDURE, w=6000))
    ch.x_axis.spPr = GraphicalProperties(ln=LineProperties(solidFill=T.GRIS_CLAIR))
    ch.y_axis.spPr = GraphicalProperties(ln=LineProperties(noFill=True))
    ch.x_axis.delete = False
    ch.y_axis.delete = False
    return ch


CAT = Reference(rc, min_col=3, min_row=L_RC_1, max_row=L_RC_N)

T.titre_section(tb, 14, 2, 13, 'LE MOIS EN QUATRE GRAPHIQUES')

ch1 = BarChart(); ch1.type = 'col'
ch1.add_data(Reference(rc, min_col=19, min_row=11, max_row=L_RC_N), titles_from_data=True)
ch1.set_categories(CAT)
ch1.series[0].graphicalProperties = _gp(T.ACIER)
ch1.gapWidth = 40
habiller(ch1, 'Achats fournisseurs, jour par jour')
tb.add_chart(ch1, 'B16')

ch2 = LineChart()
ch2.add_data(Reference(rc, min_col=20, min_row=11, max_row=L_RC_N), titles_from_data=True)
ch2.set_categories(CAT)
s = ch2.series[0]
s.graphicalProperties.line = LineProperties(solidFill=T.MARINE, w=22000)
s.marker = Marker(symbol='circle', size=5)
s.smooth = False
habiller(ch2, 'Caisse totale (PARADISE + ZENATA), soir après soir')
tb.add_chart(ch2, 'H16')

ch3 = BarChart(); ch3.type = 'col'; ch3.grouping = 'clustered'
ch3.add_data(Reference(rc, min_col=21, max_col=21, min_row=11, max_row=L_RC_N),
             titles_from_data=True)
ch3.add_data(Reference(rc, min_col=22, max_col=22, min_row=11, max_row=L_RC_N),
             titles_from_data=True)
ch3.set_categories(CAT)
ch3.series[0].graphicalProperties = _gp(T.VERT)
ch3.series[1].graphicalProperties = _gp(T.ROUGE)
ch3.gapWidth = 40
habiller(ch3, 'Encaissements et sorties, jour par jour', legende='b')
tb.add_chart(ch3, 'B34')

ch4 = BarChart(); ch4.type = 'bar'
ch4.add_data(Reference(fo, min_col=13, min_row=L_TOP_1 - 1, max_row=L_TOP_N),
             titles_from_data=True)
ch4.set_categories(Reference(fo, min_col=12, min_row=L_TOP_1, max_row=L_TOP_N))
ch4.series[0].graphicalProperties = _gp(T.MARINE)
ch4.gapWidth = 35
ch4.x_axis.tickLblSkip = 1
ch4.x_axis.tickMarkSkip = 1
habiller(ch4, 'Les dix premiers fournisseurs du mois', hauteur=8.4)
tb.add_chart(ch4, 'H34')

for r in range(15, 54):
    tb.row_dimensions[r].height = 15
T.note(tb, 54, 2, 13,
       "Tout se calcule à partir des feuilles 01 à 31 : il n'y a rien à saisir sur "
       "cette page. Les réglages du mois (année, mois, soldes d'ouverture) sont en "
       "haut de RÉCAP DU JOURNAL.", 30)
zone_impression(tb, 'A1:N54')


# =========================================================================
# 9. Données reprises de PARADISEHAMZA.xlsx  (feuille du 01/09/2026)
# =========================================================================
if REPRENDRE:
    j1 = wb['01']
    for i, (bl, frs, net, esp, chq, nchq, eff, neff) in enumerate(D.ACHATS_J1):
        r = L_ACH_1 + i
        j1.cell(r, 2).value = bl
        j1.cell(r, 3).value = frs
        j1.cell(r, 4).value = net
        j1.cell(r, 5).value = esp
        j1.cell(r, 6).value = chq
        j1.cell(r, 7).value = nchq
        j1.cell(r, 8).value = eff
        j1.cell(r, 9).value = neff

# =========================================================================
# 10. Ordre des onglets, propriétés et enregistrement
# =========================================================================
ORDRE = (['TABLEAU DE BORD', 'RÉCAP DU JOURNAL'] + JOURS +
         ['FOURNISSEURS', 'ÉCHÉANCIER', 'CHARGES FIXES', 'SUIVI MENSUEL', 'LISTES'])
wb._sheets = [wb[n] for n in ORDRE]
wb['LISTES'].sheet_state = 'hidden'
wb.active = 0

wb.properties.title = f'Caisse Paradise Aluminium — {MOIS_FR[MOIS_N]} {ANNEE}'
wb.properties.creator = 'PARADISE ALUMINIUM SARL'
wb.properties.company = 'PARADISE ALUMINIUM SARL'
wb.properties.description = (
    "Journal de caisse : fournisseurs, caisse PARADISE, caisse ZENATA, "
    "charges fixes, échéancier des chèques et des effets.")

sortie = args.sortie or os.path.join(
    RACINE, '..', f'CAISSE_PARADISE_ALUMINIUM_{MOIS_N:02d}-{ANNEE}.xlsx')
sortie = os.path.abspath(sortie)
wb.save(sortie)
print(f'✓ {sortie}')
print(f'  {len(wb.worksheets)} feuilles  ·  {MOIS_FR[MOIS_N]} {ANNEE}'
      f'  ·  caisse PARADISE {CAISSE_P:,.0f} DH  ·  caisse ZENATA {CAISSE_Z:,.0f} DH'
      .replace(',', ' '))
