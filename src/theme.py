# -*- coding: utf-8 -*-
"""Charte graphique du classeur - palette "cafe" sobre et lisible."""
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

FONT = 'Arial'

# Palette ---------------------------------------------------------------
ESPRESSO   = '3E2723'   # brun tres fonce  - bandeaux principaux
MOKA       = '5D4037'   # brun moyen       - sous-titres
CARAMEL    = 'C8A265'   # or/caramel       - accents, filets
CREME      = 'FBF6EF'   # fond clair       - zones de lecture
CREME_2    = 'F3EADD'   # fond clair alt   - lignes alternees
SABLE      = 'E8DCC8'   # bordures douces
BLANC      = 'FFFFFF'
ENCRE      = '1F1A17'   # texte principal
GRIS       = '7A6E66'   # texte secondaire
VERT       = '2E7D32'   # positif
VERT_CLAIR = 'E6F2E6'
ROUGE      = 'B3261E'   # negatif / alerte
ROUGE_CLAIR= 'FBE9E7'
BLEU       = '1A56A0'   # saisie manuelle (convention modele financier)
JAUNE      = 'FFF3C4'   # cellule a remplir
AMBRE      = 'B26A00'

# Formats de nombre -----------------------------------------------------
DH    = '#,##0 "DH";[Red]-#,##0 "DH";"–"'
DH2   = '#,##0.00 "DH";[Red]-#,##0.00 "DH";"–"'
NBR   = '#,##0;[Red]-#,##0;"–"'
PCT   = '0.0%;[Red]-0.0%;"–"'
DATE  = 'DD/MM/YYYY'
MOIS  = '[$-40C]MMMM YYYY'  # force le nom du mois en francais

# Bordures --------------------------------------------------------------
def _s(color, style='thin'):
    return Side(style=style, color=color)

BORD_LEGER  = Border(left=_s(SABLE), right=_s(SABLE), top=_s(SABLE), bottom=_s(SABLE))
BORD_BAS    = Border(bottom=_s(CARAMEL, 'medium'))
BORD_HAUT   = Border(top=_s(ESPRESSO, 'medium'))
BORD_BOITE  = Border(left=_s(CARAMEL), right=_s(CARAMEL), top=_s(CARAMEL), bottom=_s(CARAMEL))
SANS        = Border()


def police(size=10, bold=False, color=ENCRE, italic=False):
    return Font(name=FONT, size=size, bold=bold, color=color, italic=italic)


def fond(color):
    return PatternFill('solid', fgColor=color)


CENTRE  = Alignment(horizontal='center', vertical='center', wrap_text=True)
GAUCHE  = Alignment(horizontal='left',   vertical='center')
DROITE  = Alignment(horizontal='right',  vertical='center')
GAUCHE_W= Alignment(horizontal='left',   vertical='center', wrap_text=True)
HAUT_G  = Alignment(horizontal='left',   vertical='top',    wrap_text=True)


# --- Briques de mise en page ------------------------------------------
def bandeau(ws, ligne, col_debut, col_fin, titre, sous_titre=''):
    """Bandeau de titre pleine largeur en haut d'une feuille."""
    ws.merge_cells(start_row=ligne, start_column=col_debut,
                   end_row=ligne, end_column=col_fin)
    c = ws.cell(ligne, col_debut, titre)
    c.font = police(16, True, BLANC)
    c.fill = fond(ESPRESSO)
    c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws.row_dimensions[ligne].height = 38

    ws.merge_cells(start_row=ligne + 1, start_column=col_debut,
                   end_row=ligne + 1, end_column=col_fin)
    c2 = ws.cell(ligne + 1, col_debut, sous_titre)
    c2.font = police(9, False, ESPRESSO)
    c2.fill = fond(CARAMEL)
    c2.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws.row_dimensions[ligne + 1].height = 18
    for col in range(col_debut, col_fin + 1):
        ws.cell(ligne, col).fill = fond(ESPRESSO)
        ws.cell(ligne + 1, col).fill = fond(CARAMEL)


def titre_section(ws, ligne, col_debut, col_fin, texte):
    ws.merge_cells(start_row=ligne, start_column=col_debut,
                   end_row=ligne, end_column=col_fin)
    c = ws.cell(ligne, col_debut, texte)
    c.font = police(11, True, BLANC)
    c.fill = fond(MOKA)
    c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws.row_dimensions[ligne].height = 24
    for col in range(col_debut, col_fin + 1):
        ws.cell(ligne, col).fill = fond(MOKA)


def entetes(ws, ligne, col_debut, libelles, largeurs=None):
    """Ligne d'en-tetes de tableau."""
    from openpyxl.utils import get_column_letter
    for i, lib in enumerate(libelles):
        col = col_debut + i
        c = ws.cell(ligne, col, lib)
        c.font = police(9, True, BLANC)
        c.fill = fond(MOKA)
        c.alignment = CENTRE
        c.border = Border(left=_s(BLANC), right=_s(BLANC),
                          top=_s(MOKA), bottom=_s(CARAMEL, 'medium'))
        if largeurs:
            ws.column_dimensions[get_column_letter(col)].width = largeurs[i]
    ws.row_dimensions[ligne].height = 30


def ligne_total(ws, ligne, col_debut, col_fin):
    for col in range(col_debut, col_fin + 1):
        c = ws.cell(ligne, col)
        c.fill = fond(ESPRESSO)
        c.font = police(10, True, BLANC)
        c.border = Border(top=_s(CARAMEL, 'medium'))
    ws.row_dimensions[ligne].height = 24


def note(ws, ligne, col_debut, col_fin, texte, hauteur=30):
    ws.merge_cells(start_row=ligne, start_column=col_debut,
                   end_row=ligne, end_column=col_fin)
    c = ws.cell(ligne, col_debut, texte)
    c.font = police(8.5, False, GRIS, italic=True)
    c.alignment = HAUT_G
    ws.row_dimensions[ligne].height = hauteur
