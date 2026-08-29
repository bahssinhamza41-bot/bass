# -*- coding: utf-8 -*-
"""
Construit ANNEXE_ASSOCIES.xlsx.

Ce fichier ne fait PAS partie du systeme de caisse quotidien : c est une
archive. Il conserve les donnees de l ancien classeur qui n ont pas leur
place dans un classeur mensuel — les apports des deux associes et la
commande de materiel d irrigation — pour qu elles ne soient pas perdues.
"""
import os, sys
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from openpyxl import Workbook
from openpyxl.utils import get_column_letter as GL
from openpyxl.formatting.rule import CellIsRule
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.shapes import GraphicalProperties

import theme as T
import data_source as D

wb = Workbook()
wb.remove(wb.active)

asc = wb.create_sheet('ASSOCIÉS')
asc.sheet_properties.tabColor = T.MOKA
asc.sheet_view.showGridLines = False
for col, w in {'A': 4, 'B': 6, 'C': 58, 'D': 16, 'E': 16, 'F': 16, 'G': 4}.items():
    asc.column_dimensions[col].width = w

T.bandeau(asc, 1, 2, 6, 'COMPTES DES ASSOCIÉS',
          'Archive — hors système de caisse. Apports de Hamid et Mouhssine, '
          'repris ligne par ligne de la feuille DEPENSE_CAFE de l’ancien classeur.')

T.titre_section(asc, 4, 2, 6, 'SYNTHÈSE')
T.entetes(asc, 5, 2, ['', 'POSTE', 'HAMID', 'MOUHSSINE', 'TOTAL'])

AC_1 = 6
ACQ_1 = 32
ACQ_N = ACQ_1 + len(D.ACQUISITION_DETAIL) - 1
AM_1 = ACQ_N + 5
AM_N = AM_1 + len(D.AMENAGEMENT_DETAIL) - 1

lignes = [
    ('Achat du fonds de commerce (Saadaoui)', f'=$D${ACQ_N + 1}', f'=$E${ACQ_N + 1}'),
    ('Aménagement du Café Victor Hugo', f'=$D${AM_N + 1}', f'=$E${AM_N + 1}'),
    ('TOTAL APPORTÉ PAR ASSOCIÉ', f'=SUM($D${AC_1}:$D${AC_1 + 1})',
     f'=SUM($E${AC_1}:$E${AC_1 + 1})'),
    ('Part théorique due sur le fonds (795 000 DH / 2)',
     D.ACQUISITION['part_par_associe'], D.ACQUISITION['part_par_associe']),
    ('Reste à payer sur le fonds  (+ = doit,  - = a trop versé)',
     f'=$D${AC_1 + 3}-$D${AC_1}', f'=$E${AC_1 + 3}-$E${AC_1}'),
]
for i, (lib, fh, fm) in enumerate(lignes):
    r = AC_1 + i
    asc.cell(r, 3, lib)
    asc.cell(r, 4, fh)
    asc.cell(r, 5, fm)
    asc.cell(r, 6, f'=$D{r}+$E{r}')

ecarts = [
    ('Écart sur le fonds seul (Hamid - Mouhssine)', f'=$D${AC_1}-$E${AC_1}'),
    ('Écart sur l’aménagement seul (Hamid - Mouhssine)', f'=$D${AC_1 + 1}-$E${AC_1 + 1}'),
    ('ÉCART TOTAL ENTRE LES DEUX ASSOCIÉS', f'=$D${AC_1 + 2}-$E${AC_1 + 2}'),
]
for i, (lib, f) in enumerate(ecarts):
    r = AC_1 + 5 + i
    asc.cell(r, 3, lib)
    asc.cell(r, 6, f)

for i in range(8):
    rr = AC_1 + i
    gras = i in (2, 4, 7)
    for col in range(2, 7):
        cell = asc.cell(rr, col)
        cell.border = T.BORD_LEGER
        cell.fill = T.fond(T.CREME_2 if gras else (T.CREME if i % 2 else T.BLANC))
        cell.number_format = T.DH; cell.alignment = T.DROITE
        cell.font = T.police(10, gras, T.ESPRESSO if gras else T.ENCRE)
    asc.cell(rr, 3).alignment = T.GAUCHE
    asc.cell(rr, 3).number_format = 'General'
    asc.row_dimensions[rr].height = 20

asc.conditional_formatting.add(f'D{AC_1 + 4}:F{AC_1 + 7}', CellIsRule(
    operator='lessThan', formula=['0'], font=T.police(10, True, T.VERT)))
asc.conditional_formatting.add(f'D{AC_1 + 4}:F{AC_1 + 7}', CellIsRule(
    operator='greaterThan', formula=['0'], font=T.police(10, True, T.ROUGE)))

T.note(asc, AC_1 + 9, 2, 6,
       "Lecture : un « reste à payer » négatif signifie que l’associé a versé plus que sa part. "
       "Au dernier pointage, Hamid avait versé 238 DH de trop et Mouhssine 7 200 DH de trop sur "
       "l’achat du fonds, soit 6 962 DH d’écart entre eux. Sur l’aménagement du café, Mouhssine "
       "a apporté 150 779 DH de plus que Hamid, ce qui porte l’écart cumulé à 157 741 DH.", 44)


def table_detail(ws, lignes_detail, r1):
    for i, (lib, h, m) in enumerate(lignes_detail):
        r = r1 + i
        ws.cell(r, 2, i + 1).alignment = T.CENTRE
        ws.cell(r, 2).font = T.police(8, False, T.GRIS)
        ws.cell(r, 3, lib).alignment = T.GAUCHE
        ws.cell(r, 3).font = T.police(9)
        ws.cell(r, 4, h if h else None).number_format = T.DH
        ws.cell(r, 5, m if m else None).number_format = T.DH
        ws.cell(r, 6, f'=N($D{r})+N($E{r})').number_format = T.DH
        for col in range(2, 7):
            cell = ws.cell(r, col); cell.border = T.BORD_LEGER
            cell.fill = T.fond(T.CREME if i % 2 else T.BLANC)
            if col >= 4:
                cell.alignment = T.DROITE
                cell.font = T.police(9, False, T.BLEU if col < 6 else T.ENCRE)
        ws.row_dimensions[r].height = 16
    rt = r1 + len(lignes_detail)
    ws.cell(rt, 3, 'TOTAL')
    for col in (4, 5, 6):
        L = GL(col)
        ws.cell(rt, col, f'=SUM({L}{r1}:{L}{r1 + len(lignes_detail) - 1})')
        ws.cell(rt, col).number_format = T.DH_TOT
    T.ligne_total(ws, rt, 2, 6)
    for col in (4, 5, 6):
        ws.cell(rt, col).alignment = T.DROITE
    ws.cell(rt, 3).alignment = T.Alignment(horizontal='left', vertical='center', indent=1)


T.titre_section(asc, ACQ_1 - 2, 2, 6,
                'DÉTAIL — ACHAT DU FONDS DE COMMERCE (SAADAOUI, 795 000 DH)')
T.entetes(asc, ACQ_1 - 1, 2, ['N°', 'OPÉRATION', 'HAMID', 'MOUHSSINE', 'TOTAL'])
table_detail(asc, D.ACQUISITION_DETAIL, ACQ_1)

T.titre_section(asc, AM_1 - 2, 2, 6,
                'DÉTAIL — AMÉNAGEMENT ET ÉQUIPEMENT DU CAFÉ VICTOR HUGO')
T.entetes(asc, AM_1 - 1, 2, ['N°', 'OPÉRATION', 'HAMID', 'MOUHSSINE', 'TOTAL'])
table_detail(asc, D.AMENAGEMENT_DETAIL, AM_1)

T.note(asc, AM_N + 3, 2, 6,
       "Historique complet repris de la feuille DEPENSE_CAFE, ligne par ligne, sans modification. "
       "Données figées : elles ne bougent plus. Les montants en bleu sont des valeurs saisies.", 34)

g = BarChart(); g.type = 'col'; g.grouping = 'stacked'; g.overlap = 100; g.gapWidth = 60
g.add_data(Reference(asc, min_col=4, min_row=5, max_row=AC_1 + 1), titles_from_data=True)
g.add_data(Reference(asc, min_col=5, min_row=5, max_row=AC_1 + 1), titles_from_data=True)
g.set_categories(Reference(asc, min_col=3, min_row=AC_1, max_row=AC_1 + 1))
g.series[0].graphicalProperties = GraphicalProperties(solidFill='5D4037')
g.series[1].graphicalProperties = GraphicalProperties(solidFill='C8A265')
g.title = 'Apports par associé (DH)'
g.height, g.width = 8, 17
g.y_axis.numFmt = '#,##0'
asc.add_chart(g, 'B18')

# --- Commande de materiel d irrigation (feuille PVC de l ancien classeur) ---
ar = wb.create_sheet('ARCHIVE DIVERS')
ar.sheet_properties.tabColor = T.GRIS
ar.sheet_view.showGridLines = False
for col, w in {'A': 4, 'B': 40, 'C': 16, 'D': 4, 'E': 40, 'F': 16}.items():
    ar.column_dimensions[col].width = w
T.bandeau(ar, 1, 2, 6, 'ARCHIVE — COMMANDE HORS CAFÉ',
          'Commande de matériel d’irrigation reprise de la feuille « PVC ». '
          'Sans rapport avec le café, conservée pour ne rien perdre.')
chine = [('raccord droit 16 mm', 10000), ('té 16 mm', 5000), ('coude 16 mm', 5000),
         ('bouchon 16 mm', 5000), ('départ 16 mm', 2000), ('vannette 16 mm', 2000)]
turquie = [('filtre disque 2"', 100), ('filtre tamis 2"', 100), ('manomètre', 200),
           ('régulateur', 200), ('vanne', 200), ('raccords PE 32/40/50 mm', 500)]
for col0, titre, items in ((2, 'COMMANDE TEST CHINE — environ 25 000 DH', chine),
                           (5, 'COMMANDE TEST TURQUIE — environ 30 000 DH', turquie)):
    T.titre_section(ar, 4, col0, col0 + 1, titre)
    T.entetes(ar, 5, col0, ['DÉSIGNATION', 'QUANTITÉ'])
    for i, (lib, q) in enumerate(items):
        r = 6 + i
        ar.cell(r, col0, lib).alignment = T.GAUCHE
        ar.cell(r, col0 + 1, q).number_format = T.NBR
        for c in (col0, col0 + 1):
            ar.cell(r, c).border = T.BORD_LEGER
            ar.cell(r, c).fill = T.fond(T.CREME if i % 2 else T.BLANC)
            ar.cell(r, c).font = T.police(9)
        ar.cell(r, col0 + 1).alignment = T.DROITE

for ws in wb.worksheets:
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True

SORTIE = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                      'ANNEXE_ASSOCIES.xlsx')
wb.save(SORTIE)
print('Annexe ecrite :', SORTIE)
