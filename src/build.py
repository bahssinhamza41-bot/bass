# -*- coding: utf-8 -*-
"""
Construit CAISSE_VICTOR_HUGO_08-2026.xlsx : le systeme de caisse du
Cafe Victor Hugo (Mohammedia, Maroc), pour UN mois.

Structure voulue par le proprietaire :
  - une feuille par jour  (onglets 01 a 31)  <- la saisie
  - RECAP DU JOURNAL      : le mois jour par jour
  - TABLEAU DE BORD       : indicateurs et graphiques
  - CHARGES FIXES         : budget contre realise
  - SUIVI MENSUEL         : l annee mois par mois

Une feuille LISTES, masquee, ne contient que les listes deroulantes.
"""
import os, sys, datetime
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from openpyxl import Workbook
from openpyxl.utils import get_column_letter as GL
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule, DataBarRule, ColorScaleRule
from openpyxl.comments import Comment
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.marker import Marker
from openpyxl.drawing.line import LineProperties
from openpyxl.chart.shapes import GraphicalProperties

import argparse

import theme as T
import data_source as D

ap = argparse.ArgumentParser(
    description="Construit le classeur de caisse d'un mois.")
ap.add_argument('--annee', type=int, default=2026)
ap.add_argument('--mois', type=int, default=8, choices=range(1, 13), metavar='1-12')
ap.add_argument('--caisse', type=float, default=0,
                help="Solde d'especes le matin du 1er du mois "
                     "(= solde de fin du mois precedent).")
ap.add_argument('--vide', action='store_true',
                help="Ne transfere aucune donnee, meme pour aout 2026.")
args = ap.parse_args()

ANNEE, MOIS_N = args.annee, args.mois
CAISSE_DEPART = args.caisse
# Les 27 journees reprises de l ancien classeur ne concernent qu aout 2026.
TRANSFERER = (ANNEE, MOIS_N) == (2026, 8) and not args.vide

# Toujours 31 feuilles : les jours qui n existent pas dans le mois se
# neutralisent d eux-memes, ce qui permet de reutiliser le meme classeur
# pour un mois de 28, 30 ou 31 jours.
NB_JOURS = 31
JOURS = [f'{d:02d}' for d in range(1, NB_JOURS + 1)]

# --- Geometrie d une feuille jour ---------------------------------------
L_1, L_N = 12, 25          # lignes de saisie des quatre blocs
L_TOT = 26                 # ligne des totaux de bloc
L_REC = 29                 # premiere ligne du bloc recettes / solde
COL = {'achat': (2, 3), 'fixe': (5, 6), 'salaire': (8, 9), 'virement': (11, 12)}

# Libelles fixes : ils occupent toujours la meme ligne sur les 31 feuilles,
# ce qui permet de les additionner d une feuille a l autre.
CHARGES_FIXES = [
    'LOYER', 'EAU + ÉLECTRICITÉ', 'INTERNET & TÉLÉPHONE', 'TAXES & IMPÔTS',
    'ASSURANCE', 'ABONNEMENTS TV', 'ENTRETIEN & RÉPARATION', 'AUTRE CHARGE FIXE',
]
PERSONNEL = [
    'BAR MEN', 'BAR MEN ALI', 'FATIMA', 'MARWA', 'LAHCEN', 'ABDELLAH',
    'AHMED', 'ALI', 'LATIFA 1', 'LATIFA 2', 'HAMZA',
]
VIREMENTS = [
    'BANQUE', 'MOUHSSINE', 'HAMID', 'AVANCE LOYER',
    'PRÉLÈVEMENT PERSONNEL', 'AUTRE VIREMENT',
]
LIG_FIXE = {lib: L_1 + i for i, lib in enumerate(CHARGES_FIXES)}
LIG_SAL = {nom: L_1 + i for i, nom in enumerate(PERSONNEL)}
LIG_VIR = {lib: L_1 + i for i, lib in enumerate(VIREMENTS)}

# Reglages : ils vivent sur RECAP DU JOURNAL, seule page de parametrage
R_AN, R_MOIS, R_CAISSE, R_OBJ = '$C$3', '$C$4', '$C$5', '$C$6'
R_JOUR1, R_NBJ = '$F$3', '$F$4'
RCP = "'RECAP DU JOURNAL'"

wb = Workbook()
wb.remove(wb.active)


def feuille(nom, couleur, zoom=100):
    ws = wb.create_sheet(nom)
    ws.sheet_properties.tabColor = couleur
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = zoom
    return ws


def larg(ws, mapping):
    for col, w in mapping.items():
        ws.column_dimensions[col].width = w


def somme_jours(colonne, ligne):
    """Additionne la meme cellule sur les 31 feuilles jour.

    `colonne` accepte indifferemment une lettre ('F') ou un numero (6).
    """
    lettre = GL(colonne) if isinstance(colonne, int) else colonne
    return '+'.join(f"'{j}'!${lettre}${ligne}" for j in JOURS)


def sumif_jours(col_lib, col_mnt, critere):
    """Cherche un libelle dans un bloc, sur les 31 feuilles jour."""
    col_lib = GL(col_lib) if isinstance(col_lib, int) else col_lib
    col_mnt = GL(col_mnt) if isinstance(col_mnt, int) else col_mnt
    return '+'.join(
        f"SUMIF('{j}'!${col_lib}${L_1}:${col_lib}${L_N},{critere},"
        f"'{j}'!${col_mnt}${L_1}:${col_mnt}${L_N})" for j in JOURS)


# =========================================================================
# 1. LISTES  -- feuille masquee, uniquement les listes deroulantes
# =========================================================================
li = feuille('LISTES', T.GRIS)
larg(li, {'A': 26, 'B': 26, 'C': 26, 'D': 26})
li['A1'] = 'FOURNISSEURS'; li['B1'] = 'CHARGES FIXES'
li['C1'] = 'PERSONNEL'; li['D1'] = 'VIREMENTS'
for c in 'ABCD':
    li[f'{c}1'].font = T.police(10, True, T.BLANC)
    li[f'{c}1'].fill = T.fond(T.MOKA)

fournisseurs = sorted({t for _, t, _, b in D.DEPENSES if b == 'A'} |
                      {'CAFE', 'MECAFE', 'BOULANGERIE', 'FRUITS', 'LEGUMES', 'LAIT',
                       'SUCRE', 'GAZ', 'GLACE', 'DIVERS'})
for i, v in enumerate(fournisseurs):
    li.cell(2 + i, 1, v)
for i, v in enumerate(CHARGES_FIXES):
    li.cell(2 + i, 2, v)
for i, v in enumerate(PERSONNEL):
    li.cell(2 + i, 3, v)
for i, v in enumerate(VIREMENTS):
    li.cell(2 + i, 4, v)
li.cell(1, 6, "Cette feuille ne sert qu'aux listes deroulantes des feuilles jour.")
li.cell(2, 6, "Ajoutez un fournisseur en bas de la colonne A, puis etendez la plage "
              "de validation si besoin.")
for r in (1, 2):
    li.cell(r, 6).font = T.police(9, False, T.GRIS, italic=True)

L_FRS = f'LISTES!$A$2:$A${1 + len(fournisseurs)}'
L_FIX = f'LISTES!$B$2:$B${1 + len(CHARGES_FIXES)}'
L_SAL = f'LISTES!$C$2:$C${1 + len(PERSONNEL)}'
L_VIR = f'LISTES!$D$2:$D${1 + len(VIREMENTS)}'


# =========================================================================
# 2. FEUILLES JOUR  -- 01 a 31, la saisie quotidienne
# =========================================================================
BLOCS = [
    ('DÉPENSES — ACHATS DU JOUR', 'achat', L_FRS, None),
    ('CHARGES FIXES', 'fixe', L_FIX, CHARGES_FIXES),
    ('SALAIRES ET AVANCES', 'salaire', L_SAL, PERSONNEL),
    ('VIREMENTS ET DIVERS', 'virement', L_VIR, VIREMENTS),
]


def construire_jour(num):
    """Une feuille de caisse pour le jour `num` du mois."""
    ws = feuille(f'{num:02d}', T.CARAMEL)
    larg(ws, {'A': 3, 'B': 26, 'C': 14, 'D': 3, 'E': 26, 'F': 14, 'G': 3,
              'H': 26, 'I': 14, 'J': 3, 'K': 26, 'L': 14, 'M': 3})

    date_f = (f'=IF({num}>{RCP}!{R_NBJ},'
              f'"— ce jour n\'existe pas dans le mois —",'
              f'DATE({RCP}!{R_AN},{RCP}!{R_MOIS},{num}))')

    ws.merge_cells('A1:M1')
    c = ws['A1']; c.value = 'CAISSE DU JOUR'
    c.font = T.police(16, True, T.BLANC)
    c.alignment = T.Alignment(horizontal='left', vertical='center', indent=1)
    ws.row_dimensions[1].height = 36

    ws.merge_cells('A2:F2')
    c = ws['A2']; c.value = date_f
    c.font = T.police(12, True, T.ESPRESSO); c.number_format = T.JOUR_LONG
    c.alignment = T.Alignment(horizontal='left', vertical='center', indent=1)
    ws.merge_cells('G2:M2')
    c = ws['G2']; c.value = 'CAFÉ VICTOR HUGO — Mohammedia'
    c.font = T.police(10, True, T.ESPRESSO)
    c.alignment = T.Alignment(horizontal='right', vertical='center', indent=1)
    ws.row_dimensions[2].height = 22
    for col in range(1, 14):
        ws.cell(1, col).fill = T.fond(T.ESPRESSO)
        ws.cell(2, col).fill = T.fond(T.CARAMEL)

    ws.merge_cells('A4:M4')
    c = ws['A4']
    c.value = ("À REMPLIR  ▸  les quatre blocs de dépenses ci-dessous  ▸  puis, en bas, "
               "la RECETTE comptée dans la caisse.   Les cases jaunes sont les vôtres ; "
               "le reste se calcule tout seul.")
    c.font = T.police(9, True, T.ESPRESSO); c.fill = T.fond(T.CREME_2)
    c.alignment = T.Alignment(horizontal='left', vertical='center', indent=1)
    c.border = T.BORD_BOITE
    ws.row_dimensions[4].height = 26
    for col in range(1, 14):
        ws.cell(4, col).fill = T.fond(T.CREME_2)

    # --- Bandeau d indicateurs -------------------------------------------
    cartes = [
        (2, 'CAISSE DE LA VEILLE', f'=$I${L_REC}', T.MOKA),
        (5, 'RECETTE DU JOUR', f'=$C${L_REC + 3}', T.VERT),
        (8, 'TOTAL DÉPENSES', f'=$I${L_REC + 2}', T.ROUGE),
        (11, 'SOLDE CE SOIR', f'=$I${L_REC + 3}', T.ESPRESSO),
    ]
    for col, lib, f, coul in cartes:
        ws.merge_cells(start_row=6, start_column=col, end_row=6, end_column=col + 1)
        a = ws.cell(6, col, lib)
        a.font = T.police(8, True, T.GRIS)
        a.alignment = T.Alignment(horizontal='left', vertical='center', indent=1)
        ws.merge_cells(start_row=7, start_column=col, end_row=7, end_column=col + 1)
        b = ws.cell(7, col, f)
        b.font = T.police(16, True, coul); b.number_format = T.DH
        b.alignment = T.Alignment(horizontal='left', vertical='center', indent=1)
        for k in (col, col + 1):
            ws.cell(6, k).fill = T.fond(T.CREME)
            ws.cell(7, k).fill = T.fond(T.CREME)
            ws.cell(6, k).border = T.Border(top=T._s(coul, 'thick'),
                                            left=T._s(T.SABLE), right=T._s(T.SABLE))
            ws.cell(7, k).border = T.Border(left=T._s(T.SABLE), right=T._s(T.SABLE),
                                            bottom=T._s(T.SABLE))
    ws.row_dimensions[6].height = 16
    ws.row_dimensions[7].height = 26

    # --- Les quatre blocs de depenses ------------------------------------
    for titre, cle, liste, fixes in BLOCS:
        cl, cm = COL[cle]
        ws.merge_cells(start_row=10, start_column=cl, end_row=10, end_column=cm)
        t = ws.cell(10, cl, titre)
        t.font = T.police(10, True, T.BLANC); t.fill = T.fond(T.MOKA)
        t.alignment = T.Alignment(horizontal='left', vertical='center', indent=1)
        ws.cell(10, cm).fill = T.fond(T.MOKA)

        for i, lib in enumerate(('LIBELLÉ', 'MONTANT')):
            h = ws.cell(11, cl + i, lib)
            h.font = T.police(8.5, True, T.BLANC); h.fill = T.fond(T.MOKA)
            h.alignment = T.CENTRE
            h.border = T.Border(bottom=T._s(T.CARAMEL, 'medium'), left=T._s(T.BLANC))

        for r in range(L_1, L_N + 1):
            i = r - L_1
            fixe = fixes[i] if (fixes and i < len(fixes)) else None
            lc = ws.cell(r, cl, fixe)
            lc.border = T.BORD_LEGER
            lc.alignment = T.GAUCHE
            lc.font = (T.police(9, True, T.ESPRESSO) if fixe
                       else T.police(10, False, T.BLEU))
            lc.fill = T.fond(T.CREME_2 if fixe else (T.CREME if i % 2 else T.BLANC))
            mc = ws.cell(r, cm)
            mc.border = T.BORD_LEGER
            mc.number_format = T.DH; mc.alignment = T.DROITE
            mc.font = T.police(10, True, T.BLEU); mc.fill = T.fond(T.JAUNE)
            ws.row_dimensions[r].height = 17

        tt = ws.cell(L_TOT, cl, 'TOTAL')
        tt.font = T.police(9, True, T.BLANC); tt.fill = T.fond(T.ESPRESSO)
        tt.alignment = T.Alignment(horizontal='left', vertical='center', indent=1)
        tv = ws.cell(L_TOT, cm, f'=SUM({GL(cm)}{L_1}:{GL(cm)}{L_N})')
        tv.font = T.police(10, True, T.BLANC); tv.fill = T.fond(T.ESPRESSO)
        tv.number_format = T.DH_TOT; tv.alignment = T.DROITE
        tv.border = T.Border(top=T._s(T.CARAMEL, 'medium'))
        tt.border = T.Border(top=T._s(T.CARAMEL, 'medium'))
        ws.row_dimensions[L_TOT].height = 22

        if liste and not fixes:
            dv = DataValidation(type='list', formula1=f'={liste}', allow_blank=True,
                                showErrorMessage=False)
            dv.promptTitle = 'Fournisseur'
            dv.prompt = 'Choisissez dans la liste, ou tapez un nouveau nom.'
            dv.showInputMessage = True
            ws.add_data_validation(dv)
            dv.add(f'{GL(cl)}{L_1}:{GL(cl)}{L_N}')
        elif liste and fixes:
            dv = DataValidation(type='list', formula1=f'={liste}', allow_blank=True,
                                showErrorMessage=False)
            ws.add_data_validation(dv)
            dv.add(f'{GL(cl)}{L_1 + len(fixes)}:{GL(cl)}{L_N}')

        dvm = DataValidation(type='decimal', operator='greaterThanOrEqual', formula1='0',
                             allow_blank=True, showErrorMessage=True)
        dvm.errorTitle = 'Montant invalide'
        dvm.error = 'Tapez un nombre de dirhams, sans texte.'
        ws.add_data_validation(dvm)
        dvm.add(f'{GL(cm)}{L_1}:{GL(cm)}{L_N}')
    return ws


def bloc_recettes(ws, num):
    """Bas de la feuille jour : recettes encaissees et solde de caisse."""
    ws.merge_cells(start_row=L_REC - 1, start_column=2, end_row=L_REC - 1, end_column=3)
    t = ws.cell(L_REC - 1, 2, 'RECETTES DU JOUR')
    t.font = T.police(10, True, T.BLANC); t.fill = T.fond(T.VERT)
    t.alignment = T.Alignment(horizontal='left', vertical='center', indent=1)
    ws.cell(L_REC - 1, 3).fill = T.fond(T.VERT)

    ws.merge_cells(start_row=L_REC - 1, start_column=8, end_row=L_REC - 1, end_column=9)
    t = ws.cell(L_REC - 1, 8, 'SOLDE DE LA CAISSE')
    t.font = T.police(10, True, T.BLANC); t.fill = T.fond(T.ESPRESSO)
    t.alignment = T.Alignment(horizontal='left', vertical='center', indent=1)
    ws.cell(L_REC - 1, 9).fill = T.fond(T.ESPRESSO)
    ws.row_dimensions[L_REC - 1].height = 22

    recettes = [
        ('Recette caisse (espèces comptées)', None, True),
        ('Glovo — espèces', None, True),
        ('Glovo — carte', None, True),
        ('TOTAL RECETTE DU JOUR', f'=SUM($C${L_REC}:$C${L_REC + 2})', False),
    ]
    for i, (lib, f, saisie) in enumerate(recettes):
        r = L_REC + i
        lc = ws.cell(r, 2, lib)
        lc.alignment = T.GAUCHE; lc.border = T.BORD_LEGER
        vc = ws.cell(r, 3, f)
        vc.number_format = T.DH; vc.alignment = T.DROITE; vc.border = T.BORD_LEGER
        if saisie:
            lc.font = T.police(10); lc.fill = T.fond(T.BLANC)
            vc.font = T.police(11, True, T.BLEU); vc.fill = T.fond(T.JAUNE)
        else:
            lc.font = T.police(10, True, T.BLANC); lc.fill = T.fond(T.VERT)
            vc.font = T.police(12, True, T.BLANC); vc.fill = T.fond(T.VERT)
            vc.number_format = T.DH_TOT
        ws.row_dimensions[r].height = 20

    veille = (f'={RCP}!{R_CAISSE}' if num == 1
              else f"='{num - 1:02d}'!$I${L_REC + 3}")
    soldes = [
        ('Caisse de la veille', veille, T.ENCRE),
        ('Recette du jour  (+)', f'=$C${L_REC + 3}', T.VERT),
        ('Dépenses du jour  (−)', f'=$C${L_TOT}+$F${L_TOT}+$I${L_TOT}+$L${L_TOT}', T.ROUGE),
        ('SOLDE EN CAISSE CE SOIR', f'=$I${L_REC}+$I${L_REC + 1}-$I${L_REC + 2}', None),
    ]
    for i, (lib, f, coul) in enumerate(soldes):
        r = L_REC + i
        lc = ws.cell(r, 8, lib)
        lc.alignment = T.GAUCHE; lc.border = T.BORD_LEGER
        vc = ws.cell(r, 9, f)
        vc.number_format = T.DH; vc.alignment = T.DROITE; vc.border = T.BORD_LEGER
        if coul is None:
            lc.font = T.police(10, True, T.BLANC); lc.fill = T.fond(T.ESPRESSO)
            vc.font = T.police(13, True, T.BLANC); vc.fill = T.fond(T.ESPRESSO)
            vc.number_format = T.DH_TOT
        else:
            lc.font = T.police(10); lc.fill = T.fond(T.BLANC)
            vc.font = T.police(11, True, coul); vc.fill = T.fond(T.BLANC)
        if num == 1 and i == 0:
            vc.fill = T.fond(T.JAUNE); vc.font = T.police(11, True, T.BLEU)
            vc.comment = Comment(
                "Caisse au 1er du mois : cette valeur se règle sur la feuille "
                "RECAP DU JOURNAL, case C5.", 'Système', height=80, width=250)

    ws.conditional_formatting.add(f'I{L_REC + 3}', CellIsRule(
        operator='lessThan', formula=['0'], font=T.police(13, True, 'FFD6D0')))

    lig = L_REC + 5
    ws.merge_cells(start_row=lig, start_column=2, end_row=lig, end_column=12)
    n = ws.cell(lig, 2)
    n.value = ("Le solde de ce soir devient automatiquement la caisse de la veille "
               "sur la feuille du lendemain. Les libellés en gris des blocs CHARGES FIXES, "
               "SALAIRES et VIREMENTS sont fixes : ils permettent au classeur de "
               "totaliser chaque poste sur tout le mois — ne les modifiez pas.")
    n.font = T.police(8.5, False, T.GRIS, italic=True)
    n.alignment = T.HAUT_G
    ws.row_dimensions[lig].height = 28

    ws.print_area = f'A1:M{lig}'
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True


# --- Transfert des donnees d aout 2026 ----------------------------------
MAP_FIXE = {'TAXE NADAFA 2026': 'TAXES & IMPÔTS', 'EAU + ELECTRICITE': 'EAU + ÉLECTRICITÉ',
            'INTERNET': 'INTERNET & TÉLÉPHONE', 'LOYER': 'LOYER'}
MAP_VIR = {'AVANCE LOYER': 'AVANCE LOYER', 'BANQUE': 'BANQUE',
           'MOUHSSINE': 'MOUHSSINE', 'HAMID': 'HAMID'}

feuilles_jour = {}
for num in range(1, NB_JOURS + 1):
    ws = construire_jour(num)
    bloc_recettes(ws, num)
    feuilles_jour[num] = ws

nb_lignes_transferees = 0
for jour, tiers, montant, bloc in (D.DEPENSES if TRANSFERER else []):
    ws = feuilles_jour[jour]
    if bloc == 'A':
        cl, cm = COL['achat']
        r = L_1
        while ws.cell(r, cl).value:
            r += 1
        ws.cell(r, cl, tiers)
        ws.cell(r, cm, montant)
    elif bloc == 'F':
        ws.cell(LIG_FIXE[MAP_FIXE[tiers]], COL['fixe'][1], montant)
    elif bloc == 'S':
        ws.cell(LIG_SAL[tiers], COL['salaire'][1], montant)
    else:
        ws.cell(LIG_VIR[MAP_VIR[tiers]], COL['virement'][1], montant)
    nb_lignes_transferees += 1

for jour, recette in (D.RECETTES.items() if TRANSFERER else []):
    feuilles_jour[jour].cell(L_REC, 3, recette)
    nb_lignes_transferees += 1


# =========================================================================
# 3. RECAP DU JOURNAL  -- le mois jour par jour + les reglages
# =========================================================================
rc = feuille('RECAP DU JOURNAL', T.MOKA)
larg(rc, {'A': 13, 'B': 11, 'C': 14, 'D': 12, 'E': 12, 'F': 15, 'G': 13, 'H': 14,
          'I': 13, 'J': 12, 'K': 15, 'L': 15, 'M': 15, 'N': 12})

T.bandeau(rc, 1, 1, 14, 'RÉCAP DU JOURNAL',
          'Le mois entier, jour par jour. Tout est repris des feuilles 01 à 31 : '
          'aucune saisie ici, sauf les quatre réglages ci-dessous.')

reglages = [
    ('Année', ANNEE, '0'),
    ('Mois (1 à 12)', MOIS_N, '0'),
    ('Caisse au 1er du mois', CAISSE_DEPART, T.DH),
    ('Objectif recette / jour', 2200, T.DH),
]
for i, (lib, val, fmt) in enumerate(reglages):
    r = 3 + i
    rc.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    lc = rc.cell(r, 1, lib)
    lc.font = T.police(9.5, True, T.ESPRESSO)
    lc.alignment = T.Alignment(horizontal='left', vertical='center', indent=1)
    lc.fill = T.fond(T.CREME); lc.border = T.BORD_LEGER
    rc.cell(r, 2).fill = T.fond(T.CREME); rc.cell(r, 2).border = T.BORD_LEGER
    vc = rc.cell(r, 3, val)
    vc.font = T.police(10, True, T.BLEU); vc.fill = T.fond(T.JAUNE)
    vc.number_format = fmt; vc.alignment = T.CENTRE; vc.border = T.BORD_LEGER
    rc.row_dimensions[r].height = 19

rc['C4'].comment = Comment(
    "Changez ce numéro pour passer le classeur sur un autre mois : les 31 feuilles "
    "jour, ce récap et le tableau de bord suivent tout seuls. Pensez alors à effacer "
    "les montants des feuilles jour.", 'Système', height=110, width=270)

for lig, lib, f, fmt in [(3, 'Premier jour du mois', f'=DATE({R_AN},{R_MOIS},1)', T.MOIS),
                         (4, 'Nombre de jours', '=DAY(EOMONTH($F$3,0))', '0')]:
    lc = rc.cell(lig, 5, lib)
    lc.font = T.police(9.5, True, T.ESPRESSO)
    lc.alignment = T.Alignment(horizontal='left', vertical='center', indent=1)
    lc.fill = T.fond(T.CREME); lc.border = T.BORD_LEGER
    vc = rc.cell(lig, 6, f)
    vc.font = T.police(10, True); vc.number_format = fmt
    vc.alignment = T.CENTRE; vc.fill = T.fond(T.CREME); vc.border = T.BORD_LEGER

rc.merge_cells('H3:N4')
c = rc['H3']; c.value = '=$F$3'
c.font = T.police(22, True, T.ESPRESSO); c.number_format = T.MOIS
c.alignment = T.Alignment(horizontal='right', vertical='center', indent=1)

RC_H, RC_1 = 8, 9
RC_N = RC_1 + NB_JOURS - 1
RC_T, RC_R = RC_N + 1, RC_N + 2

T.entetes(rc, RC_H, 1,
          ['DATE', 'JOUR', 'RECETTE\nCAISSE', 'GLOVO\nESPÈCES', 'GLOVO\nCARTE',
           'TOTAL\nRECETTE', 'ACHATS', 'CHARGES\nFIXES', 'SALAIRES', 'VIREMENTS',
           'TOTAL\nDÉPENSES', 'RÉSULTAT\nDU JOUR', 'CAISSE\nCE SOIR', 'vs OBJ.'])

for i in range(NB_JOURS):
    r, num = RC_1 + i, i + 1
    j = f"'{num:02d}'"
    A = f'$A{r}'
    rc.cell(r, 1, f'=IF({num}>{R_NBJ},"",DATE({R_AN},{R_MOIS},{num}))')
    rc.cell(r, 2, f'=IF({A}="","",CHOOSE(WEEKDAY({A},2),"Lundi","Mardi","Mercredi",'
                  f'"Jeudi","Vendredi","Samedi","Dimanche"))')
    rc.cell(r, 3, f'=IF({A}="","",{j}!$C${L_REC})')
    rc.cell(r, 4, f'=IF({A}="","",{j}!$C${L_REC + 1})')
    rc.cell(r, 5, f'=IF({A}="","",{j}!$C${L_REC + 2})')
    rc.cell(r, 6, f'=IF({A}="","",{j}!$C${L_REC + 3})')
    rc.cell(r, 7, f'=IF({A}="","",{j}!$C${L_TOT})')
    rc.cell(r, 8, f'=IF({A}="","",{j}!$F${L_TOT})')
    rc.cell(r, 9, f'=IF({A}="","",{j}!$I${L_TOT})')
    rc.cell(r, 10, f'=IF({A}="","",{j}!$L${L_TOT})')
    rc.cell(r, 11, f'=IF({A}="","",SUM($G{r}:$J{r}))')
    rc.cell(r, 12, f'=IF({A}="","",$F{r}-$K{r})')
    rc.cell(r, 13, f'=IF({A}="","",{j}!$I${L_REC + 3})')
    rc.cell(r, 14, f'=IF({A}="","",IF($F{r}=0,"—",$F{r}-{R_OBJ}))')

    fond_l = T.CREME if i % 2 else T.BLANC
    for col in range(1, 15):
        cell = rc.cell(r, col)
        cell.border = T.BORD_LEGER; cell.fill = T.fond(fond_l)
        cell.font = T.police(10); cell.number_format = T.DH; cell.alignment = T.DROITE
    rc.cell(r, 1).number_format = T.DATE; rc.cell(r, 1).alignment = T.CENTRE
    rc.cell(r, 1).font = T.police(10, True, T.ESPRESSO)
    rc.cell(r, 2).alignment = T.CENTRE
    rc.cell(r, 2).font = T.police(9, False, T.GRIS); rc.cell(r, 2).number_format = 'General'
    rc.cell(r, 6).font = T.police(10, True, T.VERT)
    rc.cell(r, 11).font = T.police(10, True, T.ROUGE)
    rc.cell(r, 12).font = T.police(10, True)
    rc.cell(r, 13).font = T.police(10, True, T.ESPRESSO)
    rc.row_dimensions[r].height = 18

rc.cell(RC_T, 1, 'TOTAL DU MOIS')
rc.merge_cells(start_row=RC_T, start_column=1, end_row=RC_T, end_column=2)
for col in range(3, 13):
    L = GL(col)
    cell = rc.cell(RC_T, col, f'=SUM({L}{RC_1}:{L}{RC_N})')
    cell.number_format = T.DH; cell.alignment = T.DROITE
rc.cell(RC_T, 13, f'=INDEX($M${RC_1}:$M${RC_N},{R_NBJ})')
rc.cell(RC_T, 14, f'=SUMIF($F${RC_1}:$F${RC_N},">0",$N${RC_1}:$N${RC_N})')
for col in (13, 14):
    rc.cell(RC_T, col).number_format = T.DH; rc.cell(RC_T, col).alignment = T.DROITE
T.ligne_total(rc, RC_T, 1, 14)
for col in range(3, 15):
    rc.cell(RC_T, col).number_format = T.DH_TOT
rc.cell(RC_T, 1).alignment = T.Alignment(horizontal='left', vertical='center', indent=1)

rc.cell(RC_R, 1, 'PART DE LA RECETTE')
rc.merge_cells(start_row=RC_R, start_column=1, end_row=RC_R, end_column=2)
for col in (7, 8, 9, 10, 11, 12):
    L = GL(col)
    cell = rc.cell(RC_R, col, f'=IFERROR({L}{RC_T}/$F${RC_T},0)')
    cell.number_format = T.PCT; cell.alignment = T.DROITE
    cell.font = T.police(9, True, T.AMBRE)
for col in range(1, 15):
    cl = rc.cell(RC_R, col)
    cl.fill = T.fond(T.CREME_2)
    cl.border = T.Border(bottom=T._s(T.CARAMEL, 'medium'))
rc.cell(RC_R, 1).font = T.police(9, True, T.ESPRESSO)
rc.cell(RC_R, 1).alignment = T.Alignment(horizontal='left', vertical='center', indent=1)
rc.row_dimensions[RC_R].height = 20

rc.conditional_formatting.add(f'L{RC_1}:L{RC_N}', CellIsRule(
    operator='lessThan', formula=['0'], font=T.police(10, True, T.ROUGE), fill=T.fond(T.ROUGE_CLAIR)))
rc.conditional_formatting.add(f'L{RC_1}:L{RC_N}', CellIsRule(
    operator='greaterThan', formula=['0'], font=T.police(10, True, T.VERT)))
rc.conditional_formatting.add(f'M{RC_1}:M{RC_N}', CellIsRule(
    operator='lessThan', formula=['0'], font=T.police(10, True, T.ROUGE), fill=T.fond(T.ROUGE_CLAIR)))
rc.conditional_formatting.add(f'N{RC_1}:N{RC_N}', CellIsRule(
    operator='greaterThanOrEqual', formula=['0'], font=T.police(10, True, T.VERT)))
rc.conditional_formatting.add(f'N{RC_1}:N{RC_N}', CellIsRule(
    operator='lessThan', formula=['0'], font=T.police(10, True, T.ROUGE)))
rc.conditional_formatting.add(f'F{RC_1}:F{RC_N}', DataBarRule(
    start_type='num', start_value=0, end_type='percentile', end_value=98,
    color='9CCC9C', showValue=True))
rc.conditional_formatting.add(f'A{RC_1}:N{RC_N}', FormulaRule(
    formula=[f'AND($A{RC_1}<>"",WEEKDAY($A{RC_1},2)>5)'],
    fill=T.fond('F2EFE6'), stopIfTrue=False))

rc.freeze_panes = f'C{RC_1}'
T.note(rc, RC_R + 2, 1, 14,
       "Les quatre cases jaunes en haut pilotent tout le classeur. Pour passer au mois suivant : "
       "faites une copie du fichier, changez le mois, remettez « Caisse au 1er du mois » au solde "
       "de fin du mois précédent, puis effacez les montants des feuilles 01 à 31. "
       "Les lignes grisées sont les samedis et dimanches ; « vs OBJ. » compare la recette du jour "
       "à l’objectif fixé en haut.", 46)


# =========================================================================
# 4. CHARGES FIXES  -- budget contre realise, sur tout le mois
# =========================================================================
cf = feuille('CHARGES FIXES', T.AMBRE)
larg(cf, {'A': 4, 'B': 42, 'C': 15, 'D': 16, 'E': 14, 'F': 13, 'G': 4})

T.bandeau(cf, 1, 2, 6, 'CHARGES FIXES ET SALAIRES',
          'Budget mensuel (à vous de le fixer) comparé au réalisé, additionné '
          'automatiquement sur les 31 feuilles jour.')

BUDGETS = {'LOYER': 17000, 'EAU + ÉLECTRICITÉ': 4000, 'INTERNET & TÉLÉPHONE': 500,
           'TAXES & IMPÔTS': 0, 'ASSURANCE': 0, 'ABONNEMENTS TV': 500,
           'ENTRETIEN & RÉPARATION': 0, 'AUTRE CHARGE FIXE': 0}
SAL_BUDGET = {'ALI': 2400, 'LATIFA 1': 1800, 'LATIFA 2': 1800, 'ABDELLAH': 2000,
              'HAMZA': 2000, 'AHMED': 7000}

CF_H, CF_1 = 4, 5
T.entetes(cf, CF_H, 2, ['POSTE DE CHARGE', 'BUDGET / MOIS', 'RÉALISÉ DU MOIS',
                        'ÉCART', 'ÉTAT'])


def table_budget(ws, ligne_debut, lignes, colonne_montant):
    """Poste | budget saisi | realise (somme des 31 jours) | ecart | etat."""
    for i, lib in enumerate(lignes):
        r = ligne_debut + i
        budget = lignes[lib]
        ws.cell(r, 2, lib).alignment = T.GAUCHE
        ws.cell(r, 2).font = T.police(10, True, T.ESPRESSO)
        b = ws.cell(r, 3, budget)
        b.font = T.police(10, True, T.BLEU); b.fill = T.fond(T.JAUNE)
        ws.cell(r, 4, '=' + somme_jours(colonne_montant, L_1 + i))
        ws.cell(r, 5, f'=$C{r}-$D{r}')
        ws.cell(r, 6, f'=IF(N($C{r})=0,IF(N($D{r})=0,"—","HORS BUDGET"),'
                      f'IF($D{r}<=$C{r},"OK","DÉPASSÉ"))')
        for col in range(2, 7):
            cell = ws.cell(r, col); cell.border = T.BORD_LEGER
            if col != 3:
                cell.fill = T.fond(T.CREME if i % 2 else T.BLANC)
                cell.font = T.police(10)
            cell.number_format = T.DH; cell.alignment = T.DROITE
        ws.cell(r, 2).alignment = T.GAUCHE
        ws.cell(r, 2).number_format = 'General'
        ws.cell(r, 2).font = T.police(10, True, T.ESPRESSO)
        ws.cell(r, 4).font = T.police(10, True)
        ws.cell(r, 6).number_format = 'General'; ws.cell(r, 6).alignment = T.CENTRE
        ws.row_dimensions[r].height = 19
    fin = ligne_debut + len(lignes) - 1
    rt = fin + 1
    ws.cell(rt, 2, 'TOTAL')
    for col in (3, 4, 5):
        L = GL(col)
        ws.cell(rt, col, f'=SUM({L}{ligne_debut}:{L}{fin})')
    T.ligne_total(ws, rt, 2, 6)
    for col in (3, 4, 5):
        ws.cell(rt, col).number_format = T.DH_TOT; ws.cell(rt, col).alignment = T.DROITE
    ws.cell(rt, 2).alignment = T.Alignment(horizontal='left', vertical='center', indent=1)
    ws.conditional_formatting.add(f'F{ligne_debut}:F{fin}', CellIsRule(
        operator='equal', formula=['"OK"'], font=T.police(9, True, T.VERT),
        fill=T.fond(T.VERT_CLAIR)))
    ws.conditional_formatting.add(f'F{ligne_debut}:F{fin}', CellIsRule(
        operator='equal', formula=['"DÉPASSÉ"'], font=T.police(9, True, T.ROUGE),
        fill=T.fond(T.ROUGE_CLAIR)))
    ws.conditional_formatting.add(f'F{ligne_debut}:F{fin}', CellIsRule(
        operator='equal', formula=['"HORS BUDGET"'], font=T.police(8, True, T.AMBRE),
        fill=T.fond(T.JAUNE)))
    return rt


CF_TOT_FIXE = table_budget(cf, CF_1, {k: BUDGETS[k] for k in CHARGES_FIXES},
                           COL['fixe'][1])

SAL_1 = CF_TOT_FIXE + 3
T.titre_section(cf, SAL_1 - 2, 2, 6, 'SALAIRES ET AVANCES DU PERSONNEL')
T.entetes(cf, SAL_1 - 1, 2, ['PERSONNE', 'SALAIRE CONVENU', 'VERSÉ CE MOIS',
                             'RESTE À VERSER', 'ÉTAT'])
CF_TOT_SAL = table_budget(cf, SAL_1, {n: SAL_BUDGET.get(n) for n in PERSONNEL},
                          COL['salaire'][1])
for i, nom in enumerate(PERSONNEL):
    if SAL_BUDGET.get(nom) is None:
        cf.cell(SAL_1 + i, 3).comment = Comment(
            "Salaire convenu inconnu : ce nom apparaît dans les paiements d’août 2026 "
            "mais ne figurait pas dans l’ancien classeur. Renseignez-le pour suivre "
            "le reste à verser.", 'Système', height=100, width=260)

SYN_1 = CF_TOT_SAL + 3
T.titre_section(cf, SYN_1 - 1, 2, 6, 'SYNTHÈSE DU MOIS')
synthese = [
    ('Achats de marchandise', f"='RECAP DU JOURNAL'!$G${RC_T}", T.MOKA),
    ('Charges fixes', f'=$D${CF_TOT_FIXE}', T.AMBRE),
    ('Salaires', f'=$D${CF_TOT_SAL}', T.AMBRE),
    ('Virements et divers', f"='RECAP DU JOURNAL'!$J${RC_T}", T.GRIS),
    ('TOTAL DES DÉPENSES DU MOIS', f"='RECAP DU JOURNAL'!$K${RC_T}", T.ROUGE),
    ('Recette du mois', f"='RECAP DU JOURNAL'!$F${RC_T}", T.VERT),
    ('RÉSULTAT DU MOIS', f"='RECAP DU JOURNAL'!$L${RC_T}", T.ESPRESSO),
]
for i, (lib, f, coul) in enumerate(synthese):
    r = SYN_1 + i
    gras = lib.isupper()
    lc = cf.cell(r, 2, lib)
    lc.font = T.police(10, gras, T.ESPRESSO); lc.alignment = T.GAUCHE
    lc.border = T.BORD_LEGER; lc.fill = T.fond(T.CREME_2 if gras else T.BLANC)
    vc = cf.cell(r, 3, f)
    vc.number_format = T.DH; vc.alignment = T.DROITE
    vc.font = T.police(12 if gras else 10, True, coul)
    vc.border = T.BORD_LEGER; vc.fill = T.fond(T.CREME_2 if gras else T.BLANC)
    cf.row_dimensions[r].height = 20
cf.conditional_formatting.add(f'C{SYN_1 + 6}', CellIsRule(
    operator='lessThan', formula=['0'], font=T.police(12, True, T.ROUGE)))

SEUIL_1 = SYN_1 + 10
T.titre_section(cf, SEUIL_1 - 2, 2, 6, 'LE CAFÉ COUVRE-T-IL SES CHARGES ?')
NBJ_SAISIS = f"COUNTIF('RECAP DU JOURNAL'!$F${RC_1}:$F${RC_N},\">0\")"

BUD_ACHATS = SEUIL_1 - 1
cf.cell(BUD_ACHATS, 2, 'Budget achats de marchandise (par mois)').font = \
    T.police(10, True, T.ESPRESSO)
cf.cell(BUD_ACHATS, 2).alignment = T.GAUCHE
_b = cf.cell(BUD_ACHATS, 3, 20500)
_b.number_format = T.DH; _b.alignment = T.DROITE
_b.font = T.police(11, True, T.BLEU); _b.fill = T.fond(T.JAUNE)
_b.border = T.BORD_BOITE
_b.comment = Comment(
    "Ce que vous prévoyez de dépenser en marchandise sur le mois (café, eau, "
    "épicerie, boulangerie, nettoyage). La valeur par défaut, 20 500 DH, vient de "
    "l’ancien classeur (feuille CHARGE). Attention : en août 2026 le réalisé a été "
    "de 24 529 DH, nettoyage compris. Ce chiffre ne sert qu’au seuil de rentabilité "
    "ci-dessous, mais il le change beaucoup : ajustez-le à votre réalité.",
    'Système', height=140, width=290)

seuils = [
    ('Recette minimum par jour (seuil de rentabilité)',
     f'=IFERROR(($C${CF_TOT_FIXE}+$C${CF_TOT_SAL}+$C${BUD_ACHATS})/30,0)', T.ROUGE,
     "Charges fixes + salaires + achats de marchandise, en budget, divisés par "
     "30 jours. En dessous de ce chiffre de recette, la journée est déficitaire."),
    ('Recette moyenne par jour, ce mois',
     f"=IFERROR('RECAP DU JOURNAL'!$F${RC_T}/{NBJ_SAISIS},0)", T.VERT,
     "Recette du mois divisée par le nombre de jours réellement saisis."),
    ('Marge de sécurité par jour',
     f'=$C${SEUIL_1 + 1}-$C${SEUIL_1}', T.ESPRESSO,
     "Recette moyenne moins le seuil. Négatif = les charges ne sont pas couvertes."),
]
for i, (lib, f, coul, msg) in enumerate(seuils):
    r = SEUIL_1 + i
    cf.cell(r, 2, lib).font = T.police(10, True, T.ESPRESSO)
    cf.cell(r, 2).alignment = T.GAUCHE
    vc = cf.cell(r, 3, f)
    vc.number_format = T.DH; vc.alignment = T.DROITE
    vc.font = T.police(11, True, coul); vc.fill = T.fond(T.CREME_2)
    vc.border = T.BORD_BOITE
    vc.comment = Comment(msg, 'Système', height=100, width=260)
cf.conditional_formatting.add(f'C{SEUIL_1 + 2}', CellIsRule(
    operator='lessThan', formula=['0'], font=T.police(11, True, T.ROUGE)))

T.note(cf, SEUIL_1 + 4, 2, 6,
       "La colonne « RÉALISÉ » additionne, pour chaque poste, la même ligne sur les 31 feuilles jour. "
       "C’est pour cela que les libellés des blocs CHARGES FIXES et SALAIRES sont fixes sur les feuilles "
       "jour : n’en changez pas l’ordre. Les budgets en jaune viennent de l’ancien classeur (feuille CHARGE) "
       "et sont libres à modifier.", 46)

g_cf = BarChart(); g_cf.type = 'bar'; g_cf.grouping = 'clustered'; g_cf.gapWidth = 45
g_cf.add_data(Reference(cf, min_col=3, min_row=CF_H, max_row=CF_TOT_FIXE - 1),
              titles_from_data=True)
g_cf.add_data(Reference(cf, min_col=4, min_row=CF_H, max_row=CF_TOT_FIXE - 1),
              titles_from_data=True)
g_cf.set_categories(Reference(cf, min_col=2, min_row=CF_1, max_row=CF_TOT_FIXE - 1))
g_cf.series[0].graphicalProperties = GraphicalProperties(solidFill='C8A265')
g_cf.series[1].graphicalProperties = GraphicalProperties(solidFill='5D4037')
g_cf.title = 'Charges fixes : budget contre réalisé (DH)'
g_cf.height, g_cf.width = 9, 19
g_cf.y_axis.numFmt = '#,##0'
cf.add_chart(g_cf, 'H4')


# =========================================================================
# 5. SUIVI MENSUEL  -- l annee mois par mois
# =========================================================================
sm = feuille('SUIVI MENSUEL', T.MOKA)
larg(sm, {'A': 4, 'B': 17, 'C': 15, 'D': 15, 'E': 11, 'F': 15, 'G': 15,
          'H': 11, 'I': 15, 'J': 4})

T.bandeau(sm, 1, 2, 9, 'SUIVI MENSUEL',
          'L’année mois par mois. Le mois en cours se remplit tout seul ; '
          'pour les autres, recopiez les totaux quand le mois est clos.')

SM_H, SM_1 = 4, 5
T.entetes(sm, SM_H, 2,
          ['MOIS', 'RECETTE', 'ACHATS\nMARCHANDISE', '% ACHATS', 'AUTRES\nDÉPENSES',
           'TOTAL\nDÉPENSES', 'MARGE', 'RÉSULTAT'])

for i in range(12):
    r, mois = SM_1 + i, i + 1
    encours = f'MONTH($B{r})={RCP}!{R_MOIS}'
    sm.cell(r, 2, f'=DATE({RCP}!{R_AN},{mois},1)')
    sm.cell(r, 3, f"=IF({encours},{RCP}!$F${RC_T},0)")
    sm.cell(r, 4, f"=IF({encours},{RCP}!$G${RC_T},0)")
    sm.cell(r, 5, f'=IFERROR($D{r}/$C{r},0)')
    sm.cell(r, 6, f"=IF({encours},{RCP}!$K${RC_T}-{RCP}!$G${RC_T},0)")
    sm.cell(r, 7, f'=$D{r}+$F{r}')
    sm.cell(r, 8, f'=IFERROR(($C{r}-$G{r})/$C{r},0)')
    sm.cell(r, 9, f'=$C{r}-$G{r}')
    fond_l = T.CREME if i % 2 else T.BLANC
    for col in range(2, 10):
        cell = sm.cell(r, col)
        cell.border = T.BORD_LEGER; cell.fill = T.fond(fond_l)
        cell.font = T.police(10); cell.number_format = T.DH; cell.alignment = T.DROITE
    sm.cell(r, 2).number_format = T.MOIS; sm.cell(r, 2).alignment = T.GAUCHE
    sm.cell(r, 2).font = T.police(10, True, T.ESPRESSO)
    sm.cell(r, 3).font = T.police(10, True, T.VERT)
    sm.cell(r, 5).number_format = T.PCT; sm.cell(r, 5).font = T.police(10, False, T.AMBRE)
    sm.cell(r, 7).font = T.police(10, True, T.ROUGE)
    sm.cell(r, 8).number_format = T.PCT
    sm.cell(r, 9).font = T.police(10, True)
    sm.row_dimensions[r].height = 19

SM_N = SM_1 + 11
SM_T = SM_N + 1
sm.cell(SM_T, 2, 'TOTAL ANNÉE')
for col in (3, 4, 6, 7, 9):
    L = GL(col)
    sm.cell(SM_T, col, f'=SUM({L}{SM_1}:{L}{SM_N})')
sm.cell(SM_T, 5, f'=IFERROR($D{SM_T}/$C{SM_T},0)')
sm.cell(SM_T, 8, f'=IFERROR($I{SM_T}/$C{SM_T},0)')
T.ligne_total(sm, SM_T, 2, 9)
for col in range(3, 10):
    sm.cell(SM_T, col).number_format = T.DH_TOT; sm.cell(SM_T, col).alignment = T.DROITE
sm.cell(SM_T, 5).number_format = T.PCT
sm.cell(SM_T, 8).number_format = T.PCT
sm.cell(SM_T, 2).alignment = T.Alignment(horizontal='left', vertical='center', indent=1)

sm.conditional_formatting.add(f'I{SM_1}:I{SM_N}', CellIsRule(
    operator='lessThan', formula=['0'], font=T.police(10, True, T.ROUGE),
    fill=T.fond(T.ROUGE_CLAIR)))
sm.conditional_formatting.add(f'C{SM_1}:C{SM_N}', DataBarRule(
    start_type='num', start_value=0, end_type='percentile', end_value=100,
    color='9CCC9C', showValue=True))
sm.conditional_formatting.add(f'E{SM_1}:E{SM_N}', ColorScaleRule(
    start_type='num', start_value=0.2, start_color='C8E6C9',
    mid_type='num', mid_value=0.4, mid_color='FFF3C4',
    end_type='num', end_value=0.6, end_color='F5C6C0'))

# --- Rappel de l annee precedente ---------------------------------------
H_1 = SM_T + 3
T.titre_section(sm, H_1 - 1, 2, 5, 'RAPPEL 2025 — repris de l’ancien classeur')
T.entetes(sm, H_1, 2, ['MOIS', 'RECETTE', 'ACHATS', '% ACHATS'])
for i, (m, rec, ach) in enumerate(D.HISTO_2025):
    r = H_1 + 1 + i
    sm.cell(r, 2, datetime.date(2025, m, 1)).number_format = T.MOIS
    sm.cell(r, 3, rec); sm.cell(r, 4, ach)
    sm.cell(r, 5, f'=IFERROR($D{r}/$C{r},0)')
    for col in range(2, 6):
        cell = sm.cell(r, col)
        cell.border = T.BORD_LEGER; cell.fill = T.fond(T.CREME_2 if i % 2 else T.BLANC)
        cell.number_format = T.DH; cell.alignment = T.DROITE
        cell.font = T.police(9, False, T.GRIS)
    sm.cell(r, 2).number_format = T.MOIS; sm.cell(r, 2).alignment = T.GAUCHE
    sm.cell(r, 2).font = T.police(9, True, T.MOKA)
    sm.cell(r, 3).font = T.police(9, True, T.BLEU)
    sm.cell(r, 4).font = T.police(9, True, T.BLEU)
    sm.cell(r, 5).number_format = T.PCT
    sm.row_dimensions[r].height = 17
H_N = H_1 + 12
sm.cell(H_N + 1, 2, 'TOTAL 2025')
for col in (3, 4):
    L = GL(col)
    sm.cell(H_N + 1, col, f'=SUM({L}{H_1 + 1}:{L}{H_N})')
sm.cell(H_N + 1, 5, f'=IFERROR($D{H_N + 1}/$C{H_N + 1},0)')
T.ligne_total(sm, H_N + 1, 2, 5)
for col in (3, 4):
    sm.cell(H_N + 1, col).number_format = T.DH
    sm.cell(H_N + 1, col).alignment = T.DROITE
sm.cell(H_N + 1, 5).number_format = T.PCT
sm.cell(H_N + 1, 5).alignment = T.DROITE
sm.cell(H_N + 1, 2).alignment = T.Alignment(horizontal='left', vertical='center', indent=1)

g_sm = BarChart(); g_sm.type = 'col'; g_sm.gapWidth = 35
g_sm.add_data(Reference(sm, min_col=3, min_row=SM_H, max_row=SM_N), titles_from_data=True)
g_sm.add_data(Reference(sm, min_col=7, min_row=SM_H, max_row=SM_N), titles_from_data=True)
g_sm.set_categories(Reference(sm, min_col=2, min_row=SM_1, max_row=SM_N))
g_sm.series[0].graphicalProperties = GraphicalProperties(solidFill='2E7D32')
g_sm.series[1].graphicalProperties = GraphicalProperties(solidFill='B3261E')
g_sm.title = 'Recette et dépenses, mois par mois (DH)'
g_sm.height, g_sm.width = 9, 22
g_sm.y_axis.numFmt = '#,##0'
sm.add_chart(g_sm, 'K4')

T.note(sm, H_N + 3, 2, 9,
       "Ce classeur ne couvre qu’un mois : seule la ligne du mois réglé sur RÉCAP DU JOURNAL se remplit "
       "automatiquement. Quand un mois est terminé, recopiez ses totaux ici en valeur (clic droit ▸ "
       "collage spécial ▸ valeurs) pour garder l’historique de l’année. Le rappel 2025 vient de la "
       "feuille Feuil1 de l’ancien classeur ; il ne suivait que la recette et les achats.", 46)


# =========================================================================
# 6. TABLEAU DE BORD
# =========================================================================
tb = feuille('TABLEAU DE BORD', T.ESPRESSO)
for col in 'ABCDEFGHIJKLMNOP':
    tb.column_dimensions[col].width = 11.5
tb.column_dimensions['A'].width = 14

T.bandeau(tb, 1, 1, 16, 'TABLEAU DE BORD',
          'La photo du mois. Tout vient des feuilles jour, via le RÉCAP DU JOURNAL.')

tb.merge_cells('A4:F4')
c = tb['A4']; c.value = f'={RCP}!$F$3'
c.font = T.police(15, True, T.ESPRESSO); c.number_format = T.MOIS
c.alignment = T.Alignment(horizontal='left', vertical='center', indent=1)
tb.merge_cells('K4:P4')
c = tb['K4']; c.value = 'CAFÉ VICTOR HUGO — Mohammedia'
c.font = T.police(11, True, T.CARAMEL); c.alignment = T.DROITE
tb.row_dimensions[4].height = 26

REC = f'{RCP}!$F${RC_T}'
DEP = f'{RCP}!$K${RC_T}'
NBJ = f"COUNTIF({RCP}!$F${RC_1}:$F${RC_N},\">0\")"


def carte(ligne, col, libelle, formule, fmt, couleur, sous):
    tb.merge_cells(start_row=ligne, start_column=col, end_row=ligne, end_column=col + 2)
    a = tb.cell(ligne, col, libelle)
    a.font = T.police(8.5, True, T.GRIS)
    a.alignment = T.Alignment(horizontal='left', vertical='center', indent=1)
    tb.merge_cells(start_row=ligne + 1, start_column=col, end_row=ligne + 1, end_column=col + 2)
    b = tb.cell(ligne + 1, col, formule)
    b.font = T.police(18, True, couleur); b.number_format = fmt
    b.alignment = T.Alignment(horizontal='left', vertical='center', indent=1)
    tb.merge_cells(start_row=ligne + 2, start_column=col, end_row=ligne + 2, end_column=col + 2)
    d = tb.cell(ligne + 2, col, sous)
    d.font = T.police(8, False, T.GRIS)
    d.alignment = T.Alignment(horizontal='left', vertical='top', indent=1)
    haut = T.Border(top=T._s(couleur, 'thick'), left=T._s(T.SABLE), right=T._s(T.SABLE))
    mil = T.Border(left=T._s(T.SABLE), right=T._s(T.SABLE))
    bas = T.Border(left=T._s(T.SABLE), right=T._s(T.SABLE), bottom=T._s(T.SABLE))
    for k in range(col, col + 3):
        for lg, bd in ((ligne, haut), (ligne + 1, mil), (ligne + 2, bas)):
            tb.cell(lg, k).fill = T.fond(T.CREME); tb.cell(lg, k).border = bd
    tb.row_dimensions[ligne].height = 16
    tb.row_dimensions[ligne + 1].height = 28
    tb.row_dimensions[ligne + 2].height = 15


for args in [
    (6, 1, 'RECETTE DU MOIS', f'={REC}', T.DH, T.VERT, 'Total encaissé ce mois'),
    (6, 5, 'DÉPENSES DU MOIS', f'={DEP}', T.DH, T.ROUGE, 'Achats, charges, salaires, virements'),
    (6, 9, 'RÉSULTAT NET', f'={REC}-{DEP}', T.DH, T.ESPRESSO, 'Recette moins dépenses'),
    (6, 13, 'MARGE NETTE', f'=IFERROR(({REC}-{DEP})/{REC},0)', T.PCT, T.AMBRE,
     'Part de la recette qui reste'),
    (10, 1, 'RECETTE MOYENNE / JOUR', f'=IFERROR({REC}/{NBJ},0)', T.DH, T.MOKA,
     'Sur les jours réellement travaillés'),
    (10, 5, 'MEILLEURE JOURNÉE', f'=IFERROR(MAX({RCP}!$F${RC_1}:$F${RC_N}),0)', T.DH,
     T.VERT, 'Recette la plus haute du mois'),
    (10, 9, 'CAISSE EN FIN DE MOIS', f'={RCP}!$M${RC_T}', T.DH, T.ESPRESSO,
     'Solde théorique d’espèces'),
    (10, 13, 'JOURS SAISIS', f'={NBJ}', T.NBR, T.MOKA, 'Jours avec une recette enregistrée'),
]:
    carte(*args)

tb.conditional_formatting.add('I7:K7', CellIsRule(
    operator='lessThan', formula=['0'], font=T.police(18, True, T.ROUGE)))
tb.conditional_formatting.add('M7:O7', CellIsRule(
    operator='lessThan', formula=['0'], font=T.police(18, True, T.ROUGE)))

# --- Tableaux d analyse --------------------------------------------------
REP_1 = 56
T.titre_section(tb, REP_1 - 2, 1, 4, 'RÉPARTITION DES DÉPENSES')
T.entetes(tb, REP_1 - 1, 1, ['NATURE', 'MONTANT', 'PART', 'PAR JOUR'], [22, 13, 10, 12])
natures = [('Achats de marchandise', 'G'), ('Charges fixes', 'H'),
           ('Salaires et avances', 'I'), ('Virements et divers', 'J')]
for i, (lib, col_rc) in enumerate(natures):
    r = REP_1 + i
    tb.cell(r, 1, lib).font = T.police(9)
    tb.cell(r, 2, f'={RCP}!${col_rc}${RC_T}')
    tb.cell(r, 3, f'=IFERROR($B{r}/{RCP}!$K${RC_T},0)')
    tb.cell(r, 4, f'=IFERROR($B{r}/{NBJ},0)')
    for col in range(1, 5):
        cell = tb.cell(r, col); cell.border = T.BORD_LEGER
        cell.fill = T.fond(T.CREME if i % 2 else T.BLANC)
    tb.cell(r, 1).alignment = T.GAUCHE
    tb.cell(r, 2).number_format = T.DH; tb.cell(r, 2).alignment = T.DROITE
    tb.cell(r, 2).font = T.police(9, True)
    tb.cell(r, 3).number_format = T.PCT; tb.cell(r, 3).alignment = T.DROITE
    tb.cell(r, 3).font = T.police(9, False, T.AMBRE)
    tb.cell(r, 4).number_format = T.DH; tb.cell(r, 4).alignment = T.DROITE
    tb.cell(r, 4).font = T.police(9, False, T.GRIS)
REP_N = REP_1 + len(natures) - 1
rt = REP_N + 1
tb.cell(rt, 1, 'TOTAL DÉPENSES')
tb.cell(rt, 2, f'=SUM($B${REP_1}:$B${REP_N})').number_format = T.DH
tb.cell(rt, 3, f'=IFERROR($B{rt}/$B{rt},0)').number_format = T.PCT
tb.cell(rt, 4, f'=IFERROR($B{rt}/{NBJ},0)').number_format = T.DH
T.ligne_total(tb, rt, 1, 4)
for col in (2, 4):
    tb.cell(rt, col).number_format = T.DH_TOT
for col in (2, 3, 4):
    tb.cell(rt, col).alignment = T.DROITE
tb.cell(rt, 1).alignment = T.Alignment(horizontal='left', vertical='center', indent=1)

# Principaux fournisseurs : recherche du libelle sur les 31 feuilles jour
poids = {}
for _, t_, m_, bloc in D.DEPENSES:
    if bloc == 'A':
        poids[t_] = poids.get(t_, 0) + m_
top_frs = [t for t, _ in sorted(poids.items(), key=lambda x: -x[1])][:12]

FRS_1 = REP_1
T.titre_section(tb, FRS_1 - 2, 6, 9, 'PRINCIPAUX FOURNISSEURS')
T.entetes(tb, FRS_1 - 1, 6, ['FOURNISSEUR', 'MONTANT', 'PART', 'PAR JOUR'], [22, 13, 10, 12])
cl_a, cm_a = GL(COL['achat'][0]), GL(COL['achat'][1])
for i, frs in enumerate(top_frs):
    r = FRS_1 + i
    tb.cell(r, 6, frs).font = T.police(9)
    tb.cell(r, 7, '=' + sumif_jours(cl_a, cm_a, f'$F{r}'))
    tb.cell(r, 8, f'=IFERROR($G{r}/{RCP}!$G${RC_T},0)')
    tb.cell(r, 9, f'=IFERROR($G{r}/{NBJ},0)')
    for col in range(6, 10):
        cell = tb.cell(r, col); cell.border = T.BORD_LEGER
        cell.fill = T.fond(T.CREME if i % 2 else T.BLANC)
    tb.cell(r, 6).alignment = T.GAUCHE
    tb.cell(r, 7).number_format = T.DH; tb.cell(r, 7).alignment = T.DROITE
    tb.cell(r, 7).font = T.police(9, True)
    tb.cell(r, 8).number_format = T.PCT; tb.cell(r, 8).alignment = T.DROITE
    tb.cell(r, 8).font = T.police(9, False, T.AMBRE)
    tb.cell(r, 9).number_format = T.DH; tb.cell(r, 9).alignment = T.DROITE
    tb.cell(r, 9).font = T.police(9, False, T.GRIS)
FRS_N = FRS_1 + len(top_frs) - 1
tb.conditional_formatting.add(f'G{FRS_1}:G{FRS_N}', DataBarRule(
    start_type='num', start_value=0, end_type='percentile', end_value=100,
    color='8D6E63', showValue=True))
tb.conditional_formatting.add(f'B{REP_1}:B{REP_N}', DataBarRule(
    start_type='num', start_value=0, end_type='percentile', end_value=100,
    color=T.CARAMEL, showValue=True))

T.note(tb, FRS_N + 3, 1, 16,
       "Le tableau des fournisseurs cherche chaque nom dans le bloc ACHATS des 31 feuilles jour : "
       "écrivez toujours un fournisseur de la même façon (utilisez la liste déroulante) pour que le total "
       "soit juste. L’ordre des lignes a été figé sur le poids constaté en août 2026 ; les montants, eux, "
       "suivent toujours le mois en cours.", 40)


# =========================================================================
# 7. GRAPHIQUES DU TABLEAU DE BORD
# =========================================================================
PAL = ['C8A265', '5D4037', 'B26A00', '8D6E63', '1A56A0', '2E7D32',
       'A67C52', '4E6E81', 'C0873F', '6D4C41', '9E9D24', '00695C']


def habiller(ch, titre, hauteur=8.6, largeur=17.5, legende='b'):
    ch.title = titre
    ch.height, ch.width = hauteur, largeur
    ch.style = 2
    if legende:
        ch.legend.position = legende
        ch.legend.overlay = False
    else:
        ch.legend = None
    if getattr(ch, 'y_axis', None) is not None:
        ch.y_axis.majorGridlines.spPr = GraphicalProperties(
            ln=LineProperties(solidFill='E8DCC8'))
        ch.y_axis.numFmt = '#,##0'
        ch.y_axis.delete = False
    if getattr(ch, 'x_axis', None) is not None:
        ch.x_axis.delete = False
    return ch


g1 = BarChart(); g1.type = 'col'; g1.grouping = 'clustered'; g1.gapWidth = 40
g1.add_data(Reference(rc, min_col=6, min_row=RC_H, max_row=RC_N), titles_from_data=True)
g1.add_data(Reference(rc, min_col=11, min_row=RC_H, max_row=RC_N), titles_from_data=True)
g1.set_categories(Reference(rc, min_col=1, min_row=RC_1, max_row=RC_N))
g1.series[0].graphicalProperties = GraphicalProperties(solidFill='2E7D32')
g1.series[1].graphicalProperties = GraphicalProperties(solidFill='B3261E')
habiller(g1, 'Recette et dépenses, jour par jour (DH)', 8.8, 18.5)
tb.add_chart(g1, 'A16')

g2 = LineChart()
g2.add_data(Reference(rc, min_col=13, min_row=RC_H, max_row=RC_N), titles_from_data=True)
g2.set_categories(Reference(rc, min_col=1, min_row=RC_1, max_row=RC_N))
g2.series[0].graphicalProperties = GraphicalProperties(
    ln=LineProperties(solidFill='3E2723', w=28000))
g2.series[0].marker = Marker(symbol='circle', size=5)
g2.series[0].smooth = False
habiller(g2, 'Caisse en fin de journée, au fil du mois (DH)', 8.8, 15.5, legende=None)
tb.add_chart(g2, 'J16')

g3 = BarChart(); g3.type = 'bar'; g3.gapWidth = 40
g3.add_data(Reference(tb, min_col=2, min_row=REP_1 - 1, max_row=REP_N), titles_from_data=True)
g3.set_categories(Reference(tb, min_col=1, min_row=REP_1, max_row=REP_N))
pts = []
for i in range(len(natures)):
    dp = DataPoint(idx=i)
    dp.graphicalProperties = GraphicalProperties(solidFill=PAL[i % len(PAL)])
    pts.append(dp)
g3.series[0].data_points = pts
habiller(g3, 'Où part l’argent (DH)', 8.5, 17, legende=None)
tb.add_chart(g3, 'A35')

g4 = BarChart(); g4.type = 'bar'; g4.gapWidth = 40
g4.add_data(Reference(tb, min_col=7, min_row=FRS_1 - 1, max_row=FRS_N), titles_from_data=True)
g4.set_categories(Reference(tb, min_col=6, min_row=FRS_1, max_row=FRS_N))
g4.series[0].graphicalProperties = GraphicalProperties(solidFill='8D6E63')
habiller(g4, 'Principaux fournisseurs du mois (DH)', 10, 17.5, legende=None)
tb.add_chart(g4, 'J35')

for lg, txt in [(15, 'ÉVOLUTION DU MOIS'), (34, 'ANALYSE DES DÉPENSES')]:
    T.titre_section(tb, lg, 1, 16, txt)
    tb.row_dimensions[lg].height = 24


# =========================================================================
# 8. Ordre des onglets et enregistrement
# =========================================================================
ordre = (['TABLEAU DE BORD', 'RECAP DU JOURNAL'] + JOURS +
         ['CHARGES FIXES', 'SUIVI MENSUEL', 'LISTES'])
wb._sheets = [wb[n] for n in ordre]
wb['LISTES'].sheet_state = 'hidden'
wb.active = 0

for nom in ('TABLEAU DE BORD', 'RECAP DU JOURNAL', 'CHARGES FIXES', 'SUIVI MENSUEL'):
    ws = wb[nom]
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_options.horizontalCentered = True

SORTIE = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                      f'CAISSE_VICTOR_HUGO_{MOIS_N:02d}-{ANNEE}.xlsx')
wb.save(SORTIE)
print('Classeur ecrit :', SORTIE)
print('Feuilles :', len(wb.sheetnames), '| lignes transferees :', nb_lignes_transferees)
